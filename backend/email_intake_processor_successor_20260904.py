#!/usr/bin/env python
"""Bounded, idempotent PDC email-analysis worker.

Email and attachment content is untrusted business data. This module never
executes content, shells through content, changes configuration, or treats mail
as authorization. It extracts a conservative deterministic proposal and sends
that proposal to one protected Supabase RPC which owns identity matching,
review/auto-apply policy, audit and idempotency.
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import shutil
import sys
import subprocess
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

try:
    from backend.pdc_jobcard_runtime_client import RpcClient as JobcardRpcClient, validate_request as validate_jobcard_request
except ModuleNotFoundError:
    from pdc_jobcard_runtime_client import RpcClient as JobcardRpcClient, validate_request as validate_jobcard_request

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ENV = ROOT / "backend" / ".env.staging"
STAGING_PROJECT_REF = "cdsmnqxtyyoeoznmbidd"
STAGING_HOST = f"{STAGING_PROJECT_REF}.supabase.co"
EXACT_GATEWAY_INSTANCE_ID = "pdc-monitor-staging-sales-uid509-v1"
MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_EXTRACTED_CHARS = 120_000
SUPPORTED_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt",
    ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".heic",
}
WORK_TYPES = (
    "PARTS", "TINT", "HOIST", "FITTING", "BUS_4X4", "FABRICATION",
    "ELECTRICAL", "TYRE", "PIT_INSPECTION", "SUBLET",
)
JOB_CARD_WORK_KEYS = {
    "PARTS": "PARTS", "TINT": "tint", "HOIST": "hoist", "FITTING": "fitting",
    "BUS_4X4": "bus4x4", "FABRICATION": "fabrication", "ELECTRICAL": "electrical",
    "TYRE": "tyre", "PIT_INSPECTION": "pitInspection", "SUBLET": "sublet",
}
WORK_KEY_TYPES = {value: key for key, value in JOB_CARD_WORK_KEYS.items()}
WARNING_LABELS = {
    "unmatched_vehicle": "Unmatched vehicle",
    "duplicate_email": "Duplicate email",
    "missing_attachment": "Missing attachment",
    "missing_jc_number": "Missing JC number",
    "missing_jita_order": "Missing Jita order",
    "unrecognised_job_line": "Unrecognised job line",
    "conflicting_vehicle_information": "Conflicting vehicle information",
    "failed_processing": "Failed processing",
}

# A line is assigned only when one deterministic rule wins. No default stage.
CLASSIFICATION_RULES: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    ("PARTS", "parts/PO wording", (r"\bparts?\b", r"\bpurchase order\b", r"\bP\.?O\.?\b", r"\bbackorder\b", r"\bkit supplied\b")),
    ("TINT", "tint wording", (r"\btint(?:ing)?\b", r"\bwindow film\b")),
    ("HOIST", "hoist/suspension wording", (r"\bhoist\b", r"\bsuspension\b", r"\bGVM\b", r"\blift kit\b", r"\bweight upgrade\b")),
    ("BUS_4X4", "Bus 4x4 wording", (r"\bbus\s*4\s*[xX]\s*4\b", r"\bdepartment\s*138\b")),
    ("FABRICATION", "fabrication wording", (r"\bfabricat", r"\bweld", r"\bservice body\b", r"\btray\b", r"\bROPS\b", r"\bmine bar\b", r"\bjacking point")),
    ("ELECTRICAL", "electrical wording", (r"\belectrical\b", r"\bwiring?\b", r"\bUHF\b", r"\bdual batter", r"\bbrake controller\b", r"\bdriving lights?\b", r"\blight bar\b", r"\breverse alarm\b", r"\bSolis\b")),
    ("TYRE", "tyre wording", (r"\btyres?\b", r"\btires?\b", r"\bwheel alignment\b")),
    ("PIT_INSPECTION", "pit wording", (r"\bpit inspection\b", r"\bpit and weigh\b", r"\broadworthy\b")),
    ("SUBLET", "sublet/external-provider wording", (r"\bsublet\b", r"\bexternal provider\b", r"\bpaint protection\b")),
    ("FITTING", "fitting/accessory wording", (r"\bfitting\b", r"\bfit\b", r"\bbull\s*bar\b", r"\btow\s*bar\b", r"\bwinch\b", r"\bcanopy\b", r"\bsnorkel\b", r"\bseat covers?\b", r"\bfloor mats?\b", r"\bside steps?\b", r"\bnudge bar\b")),
)
EXPLICIT_PREFIXES = {
    "!PARTS": "PARTS", "!TINT": "TINT", "!HOIST": "HOIST",
    "!FIT": "FITTING", "!BUS": "BUS_4X4", "!FAB": "FABRICATION",
    "!ELEC": "ELECTRICAL", "!TYRE": "TYRE", "!PIT": "PIT_INSPECTION",
    "!SUBLET": "SUBLET",
}


@dataclass
class AttachmentEvidence:
    attachment_id: str
    filename: str
    source_hash: str
    storage_path: str
    extracted_text: str = ""
    extraction_status: str = "pending"
    extraction_error: str = ""


@dataclass
class JobLine:
    line_id: str
    original_description: str
    operation_code: str
    work_type: str | None
    assignment_reason: str
    confidence: float
    quantity: float | None = None
    estimated_duration_minutes: int | None = None
    parts_details: str = ""
    source_label: str = ""


@dataclass
class ExtractionProposal:
    extraction_version: str
    source_hash: str
    subject: str
    body_text: str
    attachment_text: str
    fields: dict[str, Any]
    job_lines: list[dict[str, Any]]
    warnings: list[str]
    warning_labels: list[str]
    extraction_hash: str = ""
    evidence: list[dict[str, Any]] = field(default_factory=list)


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        if key.strip() and key.strip() not in os.environ:
            os.environ[key.strip()] = value.strip().strip('"').strip("'")


def clean_text(value: Any, limit: int = MAX_EXTRACTED_CHARS) -> str:
    text = str(value or "").replace("\x00", " ")
    text = html.unescape(text)
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[\t ]+", " ", text)
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    return text.strip()[:limit]


def first_capture(text: str, patterns: Iterable[str], max_len: int = 160) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I | re.M)
        if match:
            return clean_text(match.group(1), max_len).strip(" .,:;\t")
    return ""


def all_captures(text: str, patterns: Iterable[str], max_len: int = 80) -> list[str]:
    values: list[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.I | re.M):
            value = clean_text(match.group(1), max_len).strip(" .,:;\t").upper()
            if value and value not in values:
                values.append(value)
    return values


def extract_fields(text: str) -> dict[str, Any]:
    vins = all_captures(text, (r"\bVIN\s*[:#-]?\s*([A-HJ-NPR-Z0-9]{17})\b", r"\b([A-HJ-NPR-Z0-9]{17})\b"), 17)
    stocks = all_captures(text, (
        r"\b(?:batch|stock|key)(?:\s*(?:no|number))?\s*[:#-]?\s*([A-Z0-9][A-Z0-9-]{4,23})\b",
        r"\b((?:IS|NS|TY|PMB|PDC)\d{5,12})\b",
    ), 24)
    orders = all_captures(text, (
        r"\b(?:Toyota\s*)?(?:order|order number)\s*[:#-]?\s*([A-Z0-9][A-Z0-9-]{4,23})\b",
    ), 24)
    jc_numbers = all_captures(text, (
        r"\b(?:job\s*card|jobcard|JC|repair\s*order)(?:\s*(?:no|number))?\s*[:#-]?\s*([A-Z0-9][A-Z0-9-]{3,23})\b",
    ), 24)
    jita = first_capture(text, (
        r"\bJITA(?:\s*(?:order|preorder|reference|no|number))?\s*[:#-]\s*([^\n]{2,80})",
        r"\bJITA\s+((?:ordered|confirmed|yes)\b[^\n]*)",
    ), 100)
    return {
        "stock_numbers": stocks,
        "toyota_order_numbers": orders,
        "vins": vins,
        "jc_number": jc_numbers[0] if len(jc_numbers) == 1 else "",
        "jc_candidates": jc_numbers,
        "jita_order": jita,
        "customer": first_capture(text, (r"^\s*(?:customer|client)\s*[:#-]\s*([^\n]{2,120})",), 120),
        "vehicle": first_capture(text, (r"^\s*(?:vehicle|model|make\s*&\s*model)\s*[:#-]\s*([^\n]{2,120})",), 120),
        "salesperson": first_capture(text, (r"^\s*(?:salesperson|sales\s*person|sales\s*rep|consultant)\s*[:#-]\s*([^\n]{2,100})",), 100),
        "eta": first_capture(text, (r"\bETA(?:\s*(?:to|at)?\s*Kewdale)?\s*[:#-]\s*([^\n]{2,80})",), 80),
        "requested_completion_date": first_capture(text, (r"\b(?:requested|required|target)\s*(?:completion|complete|delivery)\s*(?:date)?\s*[:#-]\s*([^\n]{2,80})",), 80),
        "parts_details": first_capture(text, (r"^\s*parts?(?:\s*details?)?\s*[:#-]\s*([^\n]{2,240})",), 240),
    }


def explicit_number(text: str, patterns: Iterable[str]) -> float | None:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            try:
                value = float(match.group(1))
                return value if value >= 0 else None
            except ValueError:
                return None
    return None


def classify_job_line(description: str) -> tuple[str | None, str, float]:
    description = re.sub(
        r"^\s*OP\s*[-:#/]?\s*\d{1,5}\s*[·|:—–-]*\s*", "", description,
        count=1, flags=re.I,
    )
    upper = description.upper().strip()
    current_authority = (
        (r"^SUB(?:\s|[-:#])|\bSUBLET\b", "SUBLET", "explicit SUB/Sublet evidence"),
        (r"\bWHEEL\s+NUT\s+INDICATOR", "TYRE", "wheel nut indicator authority"),
        (r"\bFIRE\s+EXTINGUISHER", "FABRICATION", "fire extinguisher authority"),
    )
    for pattern, stage, reason in current_authority:
        if re.search(pattern, description, flags=re.I):
            return stage, reason, 1.0
    explicit = [(prefix, stage) for prefix, stage in EXPLICIT_PREFIXES.items() if upper.startswith(prefix)]
    if len(explicit) == 1:
        return explicit[0][1], f"explicit source prefix {explicit[0][0]}", 1.0
    matches: list[tuple[str, str]] = []
    for stage, reason, patterns in CLASSIFICATION_RULES:
        if any(re.search(pattern, description, flags=re.I) for pattern in patterns):
            matches.append((stage, reason))
    distinct = list(dict.fromkeys(stage for stage, _ in matches))
    if len(distinct) == 1:
        reason = next(reason for stage, reason in matches if stage == distinct[0])
        return distinct[0], reason, 0.96
    if len(distinct) > 1:
        return None, "conflicting deterministic rules: " + ", ".join(distinct), 0.0
    return None, "no deterministic work-type rule matched", 0.0


def canonical_jobcard_work_key(description: str) -> str:
    """Mirror the database classifier used by non-Navision Job Card intake."""
    value = clean_text(description, 500).lower().strip(" .,;:-")
    value = re.sub(
        r"^\s*op\s*[-:#/]?\s*\d{1,5}\s*[·|:—–-]*\s*", "", value,
        count=1, flags=re.I,
    )
    cases = (
        (r"(^| )(sub|sublet)( |$)|external provider|paint protection|^!sublet", "sublet"),
        (r"wheel nut indicator|(^| )(tyres?|tires?|wheel alignment|wheel balance)( |$)|^!tyre", "tyre"),
        (r"fire extinguisher|(^| )(canopy|tray|fabricat|weld|service body|rops|mine bar|bull ?bar|jacking point)( |$)|^!fab", "fabrication"),
        (r"(^| )(uhf|radio|electrical|wiring?|spot ?lights?|driving lights?|light bar|reverse beeper|reverse alarm|whip aerial|aerial|dual batter|brake controller|solis)( |$)|^!elec", "electrical"),
        (r"(^| )(tint|tinting|window film)( |$)|^!tint", "tint"),
        (r"(^| )(hoist|suspension|gvm|lift kit|weight upgrade)( |$)|^!hoist", "hoist"),
        (r"(^| )(pit inspection|pit inspect|pit and weigh|roadworthy)( |$)|^!pit", "pitInspection"),
        (r"(^| )(parts?|purchase order|p[.]?o[.]?|backorder|kit supplied)( |$)|^!parts", "PARTS"),
        (r"(^| )(bus ?4x4|bus 4 x 4|department 138)( |$)|^!bus", "bus4x4"),
        (r"(^| )(fit|fitting|install|pre delivery|pre-delivery|long range( fuel)? tank|tow ?bar|winch|snorkel|seat covers?|floor mats?|side steps?|nudge bar|first aid|safety triangle)( |$)|^!fit", "fitting"),
    )
    return next((work_key for pattern, work_key in cases if re.search(pattern, value, flags=re.I)), "owner_supplied_document")


def source_operation_code(description: str) -> str:
    match = re.match(r"^\s*(OP\s*[-:#/]?\s*\d{1,5})\b", description, flags=re.I)
    return re.sub(r"[^A-Z0-9]", "", match.group(1).upper()) if match else ""


def operation_display_description(description: str, jc_number: str = "") -> str:
    """Remove a leading operation/JC label while retaining source text elsewhere."""
    value = clean_text(description, 500)
    value = re.sub(r"^\s*OP\s*[-:#/]?\s*\d{1,5}\s*[·|:—–-]*\s*", "", value, count=1, flags=re.I)
    if jc_number:
        value = re.sub(
            rf"^\s*(?:JC|JOB\s*CARD)(?:\s*(?:NO\.?|NUMBER))?\s*[-:#/]?\s*{re.escape(jc_number)}\s*[·|:—–-]*\s*",
            "", value, count=1, flags=re.I,
        )
    value = re.sub(r"^\s*(?:JC|JOB\s*CARD)(?:\s*(?:NO\.?|NUMBER))?\s*[-:#/]?\s*J?\d{6,15}\s*[·|:—–-]*\s*", "", value, count=1, flags=re.I)
    return value.strip() or clean_text(description, 500)


def looks_like_job_line(line: str) -> bool:
    value = clean_text(line, 500)
    if len(value) < 3 or len(value) > 500:
        return False
    if re.match(r"^(from|to|subject|date|customer|vehicle|stock|batch|vin|job\s*card|jc|order|salesperson|eta|requested|total|subtotal|gst)\b\s*[:#-]", value, re.I):
        return False
    if re.match(r"^(kind regards|regards|thanks|thank you|sent from|please see attached)\b", value, re.I):
        return False
    if re.fullmatch(r"[-_=*\s]+", value):
        return False
    return bool(
        any(re.search(pattern, value, flags=re.I) for _, _, patterns in CLASSIFICATION_RULES for pattern in patterns)
        or any(value.upper().startswith(prefix) for prefix in EXPLICIT_PREFIXES)
        or bool(source_operation_code(value))
    )


def extract_job_lines(text: str, source_label: str = "email") -> list[JobLine]:
    rows: list[JobLine] = []
    seen: set[str] = set()
    for raw in text.splitlines():
        description = clean_text(re.sub(r"^[\s•*\-]+", "", raw), 500)
        if not looks_like_job_line(description):
            continue
        normalized = re.sub(r"\s+", " ", description).strip().casefold()
        line_id = hashlib.sha256(normalized.encode("utf-8")).hexdigest()
        if line_id in seen:
            continue
        seen.add(line_id)
        stage, reason, confidence = classify_job_line(description)
        quantity = explicit_number(description, (r"\bqty\s*[:x-]?\s*(\d+(?:\.\d+)?)\b", r"\bquantity\s*[:x-]?\s*(\d+(?:\.\d+)?)\b", r"\bx\s*(\d+(?:\.\d+)?)\b"))
        hours = explicit_number(description, (r"\b(?:duration|labou?r|time)\s*[:=-]?\s*(\d+(?:\.\d+)?)\s*(?:h|hr|hrs|hours?)\b", r"\b(\d+(?:\.\d+)?)\s*(?:h|hr|hrs|hours?)\s*(?:labou?r)?\b"))
        minutes = explicit_number(description, (r"\b(?:duration|labou?r|time)\s*[:=-]?\s*(\d+)\s*(?:m|min|mins|minutes?)\b", r"\b(\d+)\s*(?:m|min|mins|minutes?)\b"))
        duration = int(round(hours * 60)) if hours is not None else (int(round(minutes)) if minutes is not None else None)
        rows.append(JobLine(
            line_id=line_id,
            original_description=description,
            operation_code=source_operation_code(description),
            work_type=stage,
            assignment_reason=reason,
            confidence=confidence,
            quantity=quantity,
            estimated_duration_minutes=duration,
            parts_details=description if stage == "PARTS" else "",
            source_label=source_label,
        ))
    return rows


def command_path(name: str, fallbacks: Iterable[str] = ()) -> str:
    found = shutil.which(name)
    if found:
        return found
    for fallback in fallbacks:
        if Path(fallback).exists():
            return fallback
    return name


def run_text_command(argv: list[str], timeout: int = 60) -> str:
    result = subprocess.run(argv, text=True, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, timeout=timeout, shell=False)
    return result.stdout if result.returncode == 0 else ""


def extract_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
    xml = re.sub(r"</w:p>", "\n", xml)
    return clean_text(re.sub(r"<[^>]+>", " ", xml))


def extract_spreadsheet(path: Path) -> str:
    try:
        from python_calamine import CalamineWorkbook  # type: ignore
        book = CalamineWorkbook.from_path(str(path))
        blocks: list[str] = []
        for name in book.sheet_names:
            sheet = book.get_sheet_by_name(name)
            blocks.append(f"[Sheet: {name}]")
            blocks.extend("\t".join(clean_text(cell, 500) for cell in row) for row in sheet.to_python(skip_empty_area=True))
        return clean_text("\n".join(blocks))
    except Exception:
        if path.suffix.lower() != ".xlsx":
            raise
    try:
        from openpyxl import load_workbook  # type: ignore
        book = load_workbook(path, read_only=True, data_only=True)
        blocks = []
        for sheet in book.worksheets:
            blocks.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                blocks.append("\t".join(clean_text(cell, 500) for cell in row))
        return clean_text("\n".join(blocks))
    except Exception:
        # Dependency-free xlsx fallback for the common inline/shared-string case.
        with zipfile.ZipFile(path) as archive:
            shared: list[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                raw = archive.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
                shared = [clean_text(re.sub(r"<[^>]+>", "", value), 500) for value in re.findall(r"<si[^>]*>(.*?)</si>", raw, re.S)]
            blocks = []
            for name in sorted(n for n in archive.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", n)):
                raw = archive.read(name).decode("utf-8", errors="replace")
                cells = []
                for attrs, value in re.findall(r"<c([^>]*)>.*?<v>(.*?)</v>.*?</c>", raw, re.S):
                    if 't="s"' in attrs and value.isdigit() and int(value) < len(shared):
                        value = shared[int(value)]
                    cells.append(clean_text(value, 500))
                blocks.append("\n".join(cells))
            return clean_text("\n".join(blocks))


def ocr_image(path: Path) -> str:
    tesseract = shutil.which("tesseract")
    if not tesseract:
        for candidate in (
            Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
            Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
        ):
            if candidate.exists():
                tesseract = str(candidate)
                break
    if tesseract:
        text = clean_text(run_text_command([tesseract, str(path), "stdout", "-l", "eng", "--psm", "6"], timeout=120))
        if text:
            return text
    try:
        from rapidocr_onnxruntime import RapidOCR
        result, _elapsed = RapidOCR()(str(path))
        return clean_text("\n".join(str(row[1]) for row in (result or []) if len(row) > 1))
    except Exception as exc:
        raise RuntimeError("No local OCR engine could read the image") from exc


def extract_pdf(path: Path) -> str:
    text = run_text_command([command_path("pdftotext"), "-layout", str(path), "-"], timeout=90)
    if clean_text(text):
        return clean_text(text)
    with tempfile.TemporaryDirectory(prefix="pdc-email-pdf-") as tmp:
        prefix = Path(tmp) / "page"
        subprocess.run([command_path("pdftoppm"), "-png", "-r", "200", str(path), str(prefix)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=180, shell=False, check=False)
        return clean_text("\n\n".join(ocr_image(image) for image in sorted(Path(tmp).glob("page-*.png"))))


def extract_attachment(path: Path) -> str:
    if not path.is_file():
        raise FileNotFoundError("attachment evidence file is missing")
    size = path.stat().st_size
    if size <= 0 or size > MAX_ATTACHMENT_BYTES:
        raise ValueError("attachment size is empty or exceeds the configured limit")
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError("attachment type is not supported")
    if ext == ".pdf":
        return extract_pdf(path)
    if ext in {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".heic"}:
        return ocr_image(path)
    if ext in {".xlsx", ".xls"}:
        return extract_spreadsheet(path)
    if ext == ".docx":
        return extract_docx(path)
    if ext == ".doc":
        return clean_text(run_text_command([command_path("antiword"), str(path)], timeout=60))
    return clean_text(path.read_text(encoding="utf-8", errors="replace"))


def analyze_record(record: dict[str, Any], attachments: list[AttachmentEvidence]) -> ExtractionProposal:
    subject = clean_text(record.get("subject"), 1000)
    body = clean_text(record.get("parsed_text") or record.get("raw_body"))
    warnings: list[str] = []
    evidence: list[dict[str, Any]] = []
    attachment_blocks: list[str] = []
    for attachment in attachments:
        path = Path(attachment.storage_path or "")
        try:
            if not attachment.source_hash:
                warnings.append("missing_attachment")
                raise ValueError("attachment hash is missing")
            if not path.is_file():
                warnings.append("missing_attachment")
                raise FileNotFoundError("attachment evidence file is missing")
            actual_hash = hashlib.sha256(path.read_bytes()).hexdigest()
            if actual_hash != attachment.source_hash:
                warnings.append("failed_processing")
                raise ValueError("attachment evidence hash mismatch")
            attachment.extracted_text = extract_attachment(path)
            attachment.extraction_status = "extracted" if attachment.extracted_text else "failed"
            if not attachment.extracted_text:
                warnings.append("failed_processing")
                attachment.extraction_error = "No text could be extracted without guessing"
            else:
                attachment_blocks.append(f"[Attachment: {attachment.filename}]\n{attachment.extracted_text}")
        except Exception as exc:
            attachment.extraction_status = "failed"
            attachment.extraction_error = str(exc)[:500]
            warnings.append("failed_processing")
        evidence.append(asdict(attachment))
    attachment_text = clean_text("\n\n".join(attachment_blocks))
    combined = clean_text("\n\n".join((subject, body, attachment_text)))
    fields = extract_fields(combined)
    job_lines: list[JobLine] = []
    job_lines.extend(extract_job_lines(body, "email body"))
    for attachment in attachments:
        if attachment.extracted_text:
            job_lines.extend(extract_job_lines(attachment.extracted_text, f"attachment:{attachment.filename}"))
    deduped: dict[str, JobLine] = {line.line_id: line for line in job_lines}
    job_lines = list(deduped.values())
    if not fields["stock_numbers"] and not fields["toyota_order_numbers"] and not fields["vins"]:
        warnings.append("unmatched_vehicle")
    if any(len(fields[key]) > 1 for key in ("stock_numbers", "toyota_order_numbers", "vins", "jc_candidates")):
        warnings.append("conflicting_vehicle_information")
    if not fields["jc_number"]:
        warnings.append("missing_jc_number")
    if any(line.work_type == "PARTS" for line in job_lines) and not fields["jita_order"]:
        warnings.append("missing_jita_order")
    if not job_lines or any(line.work_type is None for line in job_lines):
        warnings.append("unrecognised_job_line")
    warnings = list(dict.fromkeys(warnings))
    proposal = ExtractionProposal(
        extraction_version="pdc-email-intake-v1",
        source_hash=str(record.get("source_hash") or ""),
        subject=subject,
        body_text=body,
        attachment_text=attachment_text,
        fields=fields,
        job_lines=[asdict(line) for line in job_lines],
        warnings=warnings,
        warning_labels=[WARNING_LABELS[w] for w in warnings],
        evidence=evidence,
    )
    stable = asdict(proposal)
    stable.pop("extraction_hash", None)
    proposal.extraction_hash = hashlib.sha256(json.dumps(stable, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")).hexdigest()
    return proposal


def canonical_jobcard_request(record: dict[str, Any], proposal: ExtractionProposal) -> dict[str, Any]:
    """Create the exact retained pmb-email-work-v2 request, or fail closed."""
    usable = [row for row in proposal.evidence if row.get("extraction_status") == "extracted" and row.get("extracted_text")]
    if len(usable) != 1:
        raise RuntimeError("canonical job card requires exactly one extracted attachment")
    evidence = usable[0]
    source_label = f"attachment:{evidence.get('filename', '')}"
    source_lines = [line for line in proposal.job_lines if line.get("source_label") == source_label]
    if not source_lines:
        raise RuntimeError("canonical job card has no retained attachment operation lines")
    operation_lines: list[dict[str, Any]] = []
    required_work: list[str] = []
    for index, line in enumerate(source_lines, 1):
        raw_description = line.get("applied_description") or line.get("original_description")
        work_key = canonical_jobcard_work_key(str(raw_description or ""))
        minutes = line.get("estimated_duration_minutes")
        if isinstance(minutes, bool) or not isinstance(minutes, int) or minutes < 0:
            raise RuntimeError("job-card lines require explicit nonnegative source duration")
        hours = round(minutes / 60, 2)
        description = operation_display_description(raw_description, str(proposal.fields.get("jc_number") or ""))
        description = re.sub(
            r"\s+\d+(?:\.\d{1,2})?\s*(?:h|hr|hrs|hours?|m|min|mins|minutes?)\s*$",
            "", description, flags=re.I,
        )[:180].strip()
        if hours < 0 or not description:
            raise RuntimeError("job-card operation line is invalid")
        operation_no = f"OP{index}"
        operation_lines.append({
            "source_row_no": index, "operation_no": operation_no, "work_key": work_key,
            "description": description, "estimated_hours": hours,
        })
        if work_key != "owner_supplied_document" and work_key not in required_work:
            required_work.append(work_key)
    fields = proposal.fields
    stocks = list(fields.get("stock_numbers") or [])
    vins = list(fields.get("vins") or [])
    if not (len(stocks) + len(vins) == 1 and bool(fields.get("jc_number"))):
        raise RuntimeError(f"canonical job card requires one stock or VIN and one job-card number (stocks={len(stocks)}, vins={len(vins)}, jc={bool(fields.get('jc_number'))}, warnings={','.join(proposal.warnings)})")
    authentication = record.get("provider_authentication")
    if not isinstance(authentication, dict):
        raise RuntimeError("trusted provider authentication evidence is missing")
    extraction = {
        "authentication": authentication,
        "canonical_attachment_id": str(evidence.get("attachment_id") or ""),
        "canonical_document_hash": str(evidence.get("source_hash") or ""),
        "contract_version": "pmb-email-work-v2",
        "email_vehicle": {
            "cancelled": False, "conflicts": [], "customer_name": fields.get("customer") or None,
            "eta_to_kewdale": fields.get("eta") or None, "job_card_number": fields["jc_number"],
            "registration": None, "stock_numbers": stocks,
            "toyota_order_number": (fields.get("toyota_order_numbers") or [None])[0],
            "vehicle_description": fields.get("vehicle") or None, "vins": vins,
        },
        "operation_lines": operation_lines,
        "required_work": sorted(required_work),
    }
    extraction_hash = hashlib.sha256(json.dumps(extraction, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()
    return {
        "intake_id": str(record.get("id") or ""), "expected_source_hash": proposal.source_hash,
        "extraction_hash": extraction_hash,
        "provider": {
            "attachment_id": extraction["canonical_attachment_id"],
            "provider_message_id": str(record.get("internet_message_id") or record.get("graph_message_id") or ""),
            "provider_authserv_id": str(record.get("provider_authserv_id") or ""),
            "authentication": authentication,
        },
        "extraction": extraction,
    }


class SupabaseClient:
    def __init__(self, url: str, anon_key: str, access_token: str, timeout: int = 45):
        self.url = url.rstrip("/")
        self.anon_key = anon_key
        self.access_token = access_token
        self.timeout = timeout
        self.gateway_instance_id = EXACT_GATEWAY_INSTANCE_ID
        self.temp_dir = Path(tempfile.mkdtemp(prefix="pdc-monitor-attachments-"))

    def close(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _request(self, method: str, path: str, payload: Any = None) -> Any:
        data = None if payload is None else json.dumps(payload, default=str).encode("utf-8")
        request = urllib.request.Request(
            f"{self.url}/rest/v1/{path}", data=data, method=method,
            headers={"apikey": self.anon_key, "Authorization": f"Bearer {self.access_token}", "Content-Type": "application/json"},
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout) as response:
                raw = response.read(1_048_577)
                if len(raw) > 1_048_576:
                    raise RuntimeError("Supabase response exceeded 1 MiB")
                return json.loads(raw.decode("utf-8")) if raw else None
        except urllib.error.HTTPError as exc:
            raw_error = exc.read(4096)
            code = "http_error"
            try:
                body = json.loads(raw_error.decode("utf-8"))
                candidate = body.get("code") or body.get("message")
                if isinstance(candidate, str) and candidate:
                    code = re.sub(r"[^A-Za-z0-9_.:-]", "_", candidate)[:160]
            except Exception:
                pass
            raise RuntimeError(f"Supabase {method} {path} failed HTTP {exc.code} ({code})") from exc

    def rpc(self, name: str, payload: dict[str, Any]) -> dict[str, Any]:
        result = self._request("POST", f"rpc/{name}", payload)
        if not isinstance(result, dict):
            raise RuntimeError(f"RPC {name} returned invalid response")
        return result

    def pending_intakes(self, limit: int) -> list[dict[str, Any]]:
        if not 1 <= limit <= 10:
            raise RuntimeError("exact authenticated claim limit must be between 1 and 10")
        result = self.rpc("claim_pdc_email_intake_authenticated_exact_732", {"p_limit": limit, "p_gateway_instance_id": self.gateway_instance_id})
        rows = result.get("items") if result.get("ok") is True else []
        return rows if isinstance(rows, list) else []

    def _download_attachment(self, storage_path: str, filename: str, expected_hash: str = "") -> str:
        prefix = "pdc-email-attachments/"
        if not storage_path.startswith(prefix):
            legacy_prefix = "pdc-email-intake-private/"
            if not storage_path.startswith(legacy_prefix): raise RuntimeError("attachment storage path is not monitor scoped")
            bucket, object_path = legacy_prefix[:-1], storage_path[len(legacy_prefix):]
        else:
            bucket, object_path = prefix[:-1], storage_path[len(prefix):]
        request = urllib.request.Request(
            f"{self.url}/storage/v1/object/authenticated/{bucket}/{urllib.parse.quote(object_path, safe='/')}",
            headers={"apikey": self.anon_key, "Authorization": f"Bearer {self.access_token}"},
        )
        safe_suffix = Path(filename).suffix or ".bin"
        target = self.temp_dir / f"{hashlib.sha256(object_path.encode()).hexdigest()[:16]}{safe_suffix}"
        try:
            with urllib.request.urlopen(request, timeout=self.timeout) as response:
                content = response.read(MAX_ATTACHMENT_BYTES + 1)
        except urllib.error.HTTPError as exc:
            raw_error = exc.read(4096)
            code = "http_error"
            try:
                body = json.loads(raw_error.decode("utf-8")); candidate = body.get("code") or body.get("message")
                if isinstance(candidate, str) and candidate: code = re.sub(r"[^A-Za-z0-9_.:-]", "_", candidate)[:160]
            except Exception: pass
            raise RuntimeError(f"attachment download failed HTTP {exc.code} ({code})") from exc
        if len(content) > MAX_ATTACHMENT_BYTES:
            raise RuntimeError("Attachment exceeded configured size")
        if expected_hash and hashlib.sha256(content).hexdigest() != expected_hash.lower():
            raise RuntimeError("attachment download hash does not match bound source hash")
        target.write_bytes(content)
        return str(target)

    def attachments(self, intake_id: str, claim_token: str) -> list[AttachmentEvidence]:
        result = self.rpc("get_pdc_monitor_intake_attachments", {"p_intake_id": intake_id, "p_claim_token": claim_token, "p_gateway_instance_id": self.gateway_instance_id})
        rows = result.get("attachments") if result.get("ok") is True else []
        evidence: list[AttachmentEvidence] = []
        for row in rows if isinstance(rows, list) else []:
            path = self._download_attachment(str(row.get("storage_path") or ""), str(row.get("file_name") or ""), str(row.get("source_hash") or ""))
            evidence.append(AttachmentEvidence(str(row.get("id") or ""), str(row.get("file_name") or ""), str(row.get("source_hash") or ""), path))
        return evidence

    def heartbeat(self, intake_id: str, claim_token: str) -> dict[str, Any]:
        return self.rpc("heartbeat_pdc_email_intake_claim", {"p_intake_id": intake_id, "p_claim_token": claim_token, "p_gateway_instance_id": self.gateway_instance_id})

    def persist_extractions(self, intake_id: str, claim_token: str, attachments: list[AttachmentEvidence]) -> None:
        for attachment in attachments:
            if attachment.extraction_status != "extracted" or not attachment.extracted_text:
                continue
            result = self.rpc("record_pdc_monitor_attachment_extraction", {
                "p_intake_id": intake_id, "p_attachment_id": attachment.attachment_id,
                "p_claim_token": claim_token, "p_gateway_instance_id": self.gateway_instance_id,
                "p_source_hash": attachment.source_hash, "p_extracted_text": attachment.extracted_text,
            })
            if result.get("ok") is not True:
                raise RuntimeError("attachment extraction persistence failed")

    def apply_supervised_rules(self, proposal: ExtractionProposal) -> None:
        """Resolve durable active rules before canonical import; never persist a lesson."""
        for line in proposal.job_lines:
            current_key = JOB_CARD_WORK_KEYS.get(str(line.get("work_type") or ""))
            result = self.rpc("review_pdc_supervised_email_line_213", {
                "p_operation_line_id": None,
                "p_operation_code": str(line.get("operation_code") or ""),
                "p_description": str(line.get("original_description") or ""),
                "p_existing_work_key": current_key,
            })
            if result.get("ok") is not True:
                if result.get("code") == "inference_review_required":
                    line["work_type"] = None
                    line["assignment_reason"] = "persistent supervised rules require review"
                    line["confidence"] = 0.0
                    continue
                raise RuntimeError(f"supervised rule resolution failed ({str(result.get('code') or 'unknown')[:120]})")
            data = result.get("data") if isinstance(result.get("data"), dict) else {}
            work_key = str(data.get("work_key") or "")
            if work_key and work_key in WORK_KEY_TYPES:
                line["work_type"] = WORK_KEY_TYPES[work_key]
                line["assignment_reason"] = f"persistent supervised rule ({data.get('precedence') or 'rule'})"
                line["confidence"] = 1.0
                line["supervised_rule"] = {"version_id": data.get("version_id"), "version_no": data.get("version_no"), "precedence": data.get("precedence")}
                if data.get("estimated_hours") is not None:
                    hours = Decimal(str(data["estimated_hours"]))
                    if not hours.is_finite() or hours <= 0:
                        raise RuntimeError("supervised rule returned invalid estimated hours")
                    line["estimated_hours"] = hours
                    line["estimated_hours_source"] = "persistent_supervised_rule"
            line["applied_description"] = operation_display_description(
                str(line.get("original_description") or ""), str(proposal.fields.get("jc_number") or "")
            )

    def process(self, record: dict[str, Any], proposal: ExtractionProposal) -> dict[str, Any]:
        try:
            request_payload = canonical_jobcard_request(record, proposal)
        except RuntimeError as exc:
            return {
                "ok": True,
                "phase": "review",
                "code": "review_required",
                "review_reason": str(exc)[:500],
                "warnings": proposal.warnings,
                "extraction_hash": proposal.extraction_hash,
            }
        request = validate_jobcard_request(request_payload)
        provider = request["provider"]
        attestor = JobcardRpcClient(self.url, self.anon_key, self.access_token, "monitor_provider", timeout=min(180, max(1, self.timeout)))
        attested = attestor.rpc("attest_pdc_provider_email_observation", {
            "p_intake_id": request["intake_id"], "p_attachment_id": provider["attachment_id"],
            "p_expected_parent_hash": request["source_hash"], "p_expected_attachment_hash": request["attachment_hash"],
            "p_provider_message_id": provider["provider_message_id"], "p_provider_authserv_id": provider["provider_authserv_id"],
            "p_authentication": provider["authentication"],
        })
        if not isinstance(attested, dict) or attested.get("ok") is not True:
            return {"ok": False, "phase": "provider_attestation", "code": "attestation_failed"}
        return self.rpc("process_claimed_pdc_email_intake_work", {
            "p_intake_id": request["intake_id"], "p_claim_token": str(record.get("claim_token") or ""),
            "p_gateway_instance_id": self.gateway_instance_id, "p_expected_source_hash": request["source_hash"],
            "p_extraction_hash": request["extraction_hash"], "p_extraction": request["extraction"],
        })

    def record_result(self, intake_id: str, claim_token: str, success: bool, result: dict[str, Any], error_code: str = "", error: str = "", temporary: bool = False) -> dict[str, Any]:
        return self.rpc("record_pdc_email_intake_result", {
            "p_intake_id": intake_id, "p_claim_token": claim_token, "p_gateway_instance_id": self.gateway_instance_id,
            "p_success": success, "p_result": result,
            "p_error_code": error_code or None, "p_error_detail": error or None,
            "p_temporary": temporary, "p_revision_summary": result.get("revision_summary", {}) if isinstance(result, dict) else {},
        })


def run_once(client: SupabaseClient, limit: int = 20) -> dict[str, Any]:
    started = client.rpc("record_pdc_email_monitor_cycle", {"p_running_status": "running", "p_error_code": None, "p_error": None})
    if started.get("ok") is not True:
        raise RuntimeError("monitor cycle start telemetry failed")
    summary: dict[str, Any] = {"ok": True, "seen": 0, "processed": 0, "review": 0, "duplicates": 0, "failed": 0, "results": []}
    transient_codes = {"database_unavailable", "temporary_failure", "processing_failed", "backend_unavailable", "timeout"}
    try:
        records = client.pending_intakes(limit)
    except Exception as exc:
        error = f"{type(exc).__name__}: {exc}"[:8000]
        recorded = client.rpc("record_pdc_email_monitor_cycle", {
            "p_running_status": "degraded",
            "p_error_code": "queue_lookup_failed",
            "p_error": error,
        })
        if recorded.get("ok") is not True:
            raise RuntimeError(f"monitor queue lookup failed and finish telemetry failed: {error}") from exc
        raise
    for record in records:
        summary["seen"] += 1
        intake_id = str(record["id"])
        claim_token = str(record.get("claim_token") or "")
        try:
            attachments = client.attachments(intake_id, claim_token)
            proposal = analyze_record(record, attachments)
            client.apply_supervised_rules(proposal)
            client.heartbeat(intake_id, claim_token)
            client.persist_extractions(intake_id, claim_token, attachments)
            result = client.process(record, proposal)
            code = str(result.get("code") or "")
            success = result.get("ok") is True
            client.record_result(intake_id, claim_token, success, result, code, "" if success else json.dumps(result, default=str)[:4000], code in transient_codes)
            if success:
                summary["processed"] += 1
                if code in {"review_required", "processed_review"}: summary["review"] += 1
                if code in {"duplicate", "already_processed"}: summary["duplicates"] += 1
            else:
                summary["failed"] += 1; summary["ok"] = False
            summary["results"].append({"intake_id": intake_id, "code": code, "ok": success})
        except Exception as exc:
            summary["failed"] += 1; summary["ok"] = False
            error = f"{type(exc).__name__}: {exc}"[:4000]
            try: client.record_result(intake_id, claim_token, False, {}, "worker_exception", error, isinstance(exc, (urllib.error.URLError, TimeoutError)))
            except Exception as record_exc: error += f"; result_record_error={type(record_exc).__name__}: {record_exc}"
            summary["results"].append({"intake_id": intake_id, "code": "failed_processing", "ok": False, "error": error[:500]})
    final_state = "idle" if summary["ok"] else "degraded"
    error_detail = "" if summary["ok"] else json.dumps(summary["results"][-5:], default=str)[:8000]
    recorded = client.rpc("record_pdc_email_monitor_cycle", {
        "p_running_status": final_state,
        "p_error_code": None if summary["ok"] else "email_processing_failed",
        "p_error": None if summary["ok"] else error_detail,
    })
    if recorded.get("ok") is not True:
        raise RuntimeError("monitor cycle finish telemetry failed")
    return summary

def _monitor_access_token(url: str, anon_key: str) -> str:
    direct = os.environ.get("PDC_MONITOR_ACCESS_TOKEN", "").strip()
    if direct: return direct
    email_address = os.environ.get("PDC_MONITOR_EMAIL", "").strip(); password = os.environ.get("PDC_MONITOR_PASSWORD", "")
    if not email_address or not password: raise RuntimeError("Scoped PDC Monitor Viewer credentials are unavailable")
    request = urllib.request.Request(f"{url.rstrip('/')}/auth/v1/token?grant_type=password", data=json.dumps({"email": email_address, "password": password}).encode("utf-8"), method="POST", headers={"apikey": anon_key, "Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(request, timeout=30) as response: result=json.loads(response.read(65537).decode("utf-8"))
    except (urllib.error.HTTPError, urllib.error.URLError, json.JSONDecodeError) as exc: raise RuntimeError("PDC Monitor Viewer authentication failed") from exc
    token=result.get("access_token") if isinstance(result,dict) else None
    if not isinstance(token,str) or len(token)<8 or token==anon_key: raise RuntimeError("PDC Monitor Viewer authentication returned no scoped token")
    return token


def _is_exact_staging_url(value: str) -> bool:
    try:
        parsed = urllib.parse.urlsplit(value)
        port = parsed.port
    except (TypeError, ValueError):
        return False
    return (parsed.scheme == "https" and parsed.hostname == STAGING_HOST and port is None
            and parsed.username is None and parsed.password is None
            and parsed.path in ("", "/") and not parsed.query and not parsed.fragment)


def main() -> int:
    parser = argparse.ArgumentParser(description="Process retained PDC email intake into idempotent work-item proposals")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV))
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--fixture", help="Analyze one JSON fixture without network writes")
    args = parser.parse_args()
    if args.fixture:
        fixture = json.loads(Path(args.fixture).read_text(encoding="utf-8"))
        attachments = [AttachmentEvidence(**item) for item in fixture.get("attachments", [])]
        print(json.dumps(asdict(analyze_record(fixture, attachments)), indent=2, default=str))
        return 0
    load_dotenv(Path(args.env_file))
    url = os.environ.get("SUPABASE_URL", os.environ.get("PDC_STAGING_SUPABASE_URL", "")).strip()
    anon_key = os.environ.get("SUPABASE_ANON_KEY", os.environ.get("PDC_STAGING_ANON_KEY", "")).strip()
    if not url or not anon_key:
        print(json.dumps({"ok": False, "error": "STAGING SUPABASE_URL and SUPABASE_ANON_KEY are required in the ignored backend/.env.staging"}))
        return 2
    if not _is_exact_staging_url(url):
        print(json.dumps({"ok": False, "error": "Refusing non-staging Supabase project", "required_project_ref": STAGING_PROJECT_REF}))
        return 3
    try:
        client = SupabaseClient(url, anon_key, _monitor_access_token(url, anon_key))
        summary = run_once(client, max(1, args.limit))
    finally:
        if 'client' in locals(): client.close()
    print(json.dumps(summary, sort_keys=True))
    return 0 if summary["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
