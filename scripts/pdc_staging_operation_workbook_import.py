"""Controlled staging-only import for Hermes_Operation_Lines_139.xlsx.

The controller keeps unknown work mappings quarantined, temporarily pivots the
approved pmb-auditor staging identity viewer -> importer, enrolls it as a writer,
then restores viewer and revokes writer authority in a finally block. No Monitor,
mailbox, cron, Production, booking, completion, Parts, or location authority is enabled.
"""
from __future__ import annotations
import argparse, base64, collections, hashlib, json, re, urllib.error, urllib.parse, urllib.request
from pathlib import Path
from openpyxl import load_workbook
from pdc_staging_management_migration import STAGING_REF, PRODUCTION_REF, _post

WORKBOOK=Path(r"C:/Users/nwmgr/AppData/Local/hermes/cache/documents/doc_a823986dc288_Hermes_Operation_Lines_139.xlsx")
WEBSITE_ENV=Path(r"C:/Users/nwmgr/AppData/Local/hermes/profiles/website-development-lead/.env")
AUDITOR_ENV=Path(r"C:/Users/nwmgr/AppData/Local/hermes/profiles/pmb-auditor/.env")
EXPECTED_WORKBOOK_SHA="56e04ce63beb40cb539db28df1262f1cdc4166e4f08b117025162e9228719ad2"
VALID_KEYS={"bus4x4","tint","hoist","fitting","fabrication","electrical","tyre","pitInspection","PARTS","sublet"}

def env(path:Path)->dict[str,str]:
 out={}
 for raw in path.read_text(encoding="utf-8-sig").splitlines():
  s=raw.strip()
  if s and not s.startswith("#") and "=" in s:
   k,v=s.split("=",1);out[k.strip()]=v.strip().strip('"').strip("'")
 return out

def classify(description:str)->str:
 d=re.sub(r"\s+"," ",description).strip(" .,;:-").lower();dn=" ".join(description.lower().replace("-"," ").split())
 # Craig-approved owner rules precede generic tray/canopy/fit tokens.
 if re.search(r"\btow\s*bars?\b",d):return "fitting"
 if re.search(r"\b(?:long\s+range(?:r)?\s+(?:fuel\s+)?tanks?|arb\s+frontier\b.{0,80}\b(?:fuel\s+)?tanks?|sub\s+tank\s+replacem\w*)\b",d):return "hoist"
 if re.search(r"\bfire\s+ext(?:inguisher|inuisher|inguishers?|inuishers?)?\b",d):return "fitting"
 if (re.search(r"\banderson\s+plugs?\b",d)
   or re.search(r"\b12v\b.{0,60}\b(?:acc(?:essory)?\s+socket|plugs?)\b|\b(?:acc(?:essory)?\s+socket|plugs?)\b.{0,60}\b12v\b",d)
   or re.search(r"\b(?:arb\s+)?battery\s+box\b|\bbcdc\d*\b|\bxrs\s*370c\b|\bnavman\b|\bcardex\b",d)):return "electrical"
 key=None
 for pat,value in [
  (r"(^| )(uhf|radio|electrical|spot ?lights?|light bar)( |$)","electrical"),(r"(^| )(tyres?|tires?)( |$)","tyre"),
  (r"(^| )(canopy|tray|fabricat)","fabrication"),(r"(^| )(tint)( |$)","tint"),(r"(^| )(hoist)( |$)","hoist"),
  (r"(^| )(pit inspection|pit inspect)( |$)","pitInspection"),(r"(^| )(parts?)( |$)","PARTS"),(r"(^| )(bus ?4x4)( |$)","bus4x4"),
  (r"(^| )(fit|install)( |$)","fitting")]:
  if re.search(pat,d):key=value;break
 if "reflective stripe" in dn:key="sublet"
 if any(x in dn for x in ("bullbar","bull bar","bulbar")):key="fitting"
 return key or "UNMAPPED"

def payload()->tuple[list[dict],dict]:
 sha=hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
 if sha!=EXPECTED_WORKBOOK_SHA:raise RuntimeError("PDC_WORKBOOK_SHA_MISMATCH")
 wb=load_workbook(WORKBOOK,read_only=True,data_only=True);ws=wb["Hermes Import"];groups=collections.OrderedDict();rows=0
 for r in ws.iter_rows(min_row=2,values_only=True):
  if not any(v is not None for v in r):continue
  rows+=1;jc=str(r[1] or "").strip().upper();stock=str(r[5] or "").strip();reg=str(r[6] or "").strip().upper() or None
  op={"operation_no":"OP"+str(int(r[2])),"work_key":classify(str(r[3] or "").strip()),"description":str(r[3] or "").strip(),"estimated_hours":None if r[4] is None else float(r[4]),"estimated_hours_source":None if r[4] is None else "job_card"}
  groups.setdefault((jc,stock,reg),[]).append(op)
 p=[{"pair_no":n,"job_card_number":jc,"stock_number":stock,"registration":reg,"operations":ops} for n,((jc,stock,reg),ops) in enumerate(groups.items(),1)]
 counts=collections.Counter(o["work_key"] for x in p for o in x["operations"])
 if len(p)!=166 or rows!=1488 or counts["UNMAPPED"]!=983 or rows-counts["UNMAPPED"]!=505:raise RuntimeError("PDC_WORKBOOK_CLASSIFICATION_COUNT_DRIFT")
 if any(len(x["operations"])>100 for x in p) or any(len(o["description"])>180 for x in p for o in x["operations"]):raise RuntimeError("PDC_WORKBOOK_CONTRACT_LIMIT")
 return p,{"workbook_sha256":sha,"pair_count":len(p),"operation_count":rows,"mapped_operation_count":505,"quarantined_operation_count":983,"work_key_counts":dict(sorted(counts.items())),"unknown_hours":sum(o["estimated_hours"] is None for x in p for o in x["operations"]),"explicit_zero_hours":sum(o["estimated_hours"]==0 for x in p for o in x["operations"])}

def request_json(url:str,method="GET",headers=None,body=None):
 data=None if body is None else json.dumps(body,separators=(",",":")).encode()
 req=urllib.request.Request(url,data=data,method=method,headers=headers or {})
 try:
  with urllib.request.urlopen(req,timeout=300) as r:return json.load(r)
 except urllib.error.HTTPError as e:
  text=e.read().decode(errors="replace")[:800]
  raise RuntimeError(f"PDC_HTTP_{e.code}:{text}") from None

def auth(base,anon,email,password):
 s=request_json(base+"/auth/v1/token?grant_type=password","POST",{"apikey":anon,"Content-Type":"application/json"},{"email":email,"password":password})
 token=s["access_token"];uid=s["user"]["id"];h={"apikey":anon,"Authorization":"Bearer "+token,"Content-Type":"application/json"}
 roles=request_json(base+f"/rest/v1/pdc_user_roles?auth_user_id=eq.{uid}&select=role,active,account_status",headers=h)
 if len(roles)!=1 or not roles[0]["active"] or roles[0]["account_status"]!="approved":raise RuntimeError("PDC_IDENTITY_NOT_APPROVED")
 return {"token":token,"uid":uid,"email":email,"headers":h,"role":roles[0]["role"]}

def rpc(base,actor,name,body):
 r=request_json(base+"/rest/v1/rpc/"+name,"POST",actor["headers"],body)
 if not isinstance(r,dict) or not r.get("ok"):raise RuntimeError("PDC_RPC_"+name+":"+str(r.get("code") if isinstance(r,dict) else "INVALID"))
 return r

def role_change(base,admin,target_email,role,reason):
 r=request_json(base+"/rest/v1/rpc/admin_change_role","POST",admin["headers"],{"p_target_email":target_email,"p_role":role,"p_reason":reason})
 if not isinstance(r,dict) or r.get("role")!=role or not r.get("active") or r.get("account_status")!="approved":raise RuntimeError("PDC_ROLE_CHANGE_READBACK_FAILED")
 return r

def read_pairs(base,actor,preview_id,count):
 out=[]
 for offset in range(0,count,100):
  r=rpc(base,actor,"read_pdc_pmb_workbook_pair_verification",{"p_preview_id":preview_id,"p_offset":offset,"p_limit":min(100,count-offset)})
  out.extend(r["data"]["pairs"])
 if len(out)!=count:raise RuntimeError("PDC_PAIR_READ_COUNT_MISMATCH")
 return sorted(out,key=lambda x:x["pair_no"])

def server_hash(value):
 raw=json.dumps(value,separators=(",",":"),ensure_ascii=False).encode();b64=base64.b64encode(raw).decode()
 sql=f"SET TRANSACTION READ ONLY;select encode(extensions.digest(convert_to(convert_from(decode('{b64}','base64'),'UTF8')::jsonb::text,'UTF8'),'sha256'),'hex') h;"
 return _post(f"https://api.supabase.com/v1/projects/{STAGING_REF}/database/query/read-only",sql)[0]["h"]

def live_readback():
 sql="""SET TRANSACTION READ ONLY;select jsonb_build_object(
 'production_sentinel_absent',to_regclass('public.pdc_production_environment_sentinel') is null,
 'vehicles',(select count(*) from public.vehicles),
 'unique_stocks',(select count(distinct upper(btrim(stock_number))) from public.vehicles),
 'visible_vehicles',(select count(*) from public.vehicles where visible_on_board and deleted_at is null and lifecycle_state='active'),
 'operation_lines',(select count(*) from public.pdc_authenticated_email_operation_lines),
 'mapped_review_lines',(select count(*) from public.pdc_pmb_workbook_operation_reviews where disposition='accepted'),
 'quarantined_review_lines',(select count(*) from public.pdc_pmb_workbook_operation_reviews where disposition='quarantined'),
 'pair_reviews',(select count(*) from public.pdc_pmb_workbook_pair_reviews),
 'pair_approvals',(select count(*) from public.pdc_pmb_workbook_pair_approvals),
 'pair_receipts',(select count(*) from public.pdc_pmb_workbook_pair_receipts),
 'apply_receipts',(select count(*) from public.pdc_pmb_workbook_apply_receipts),
 'work_items',(select count(*) from public.vehicle_work_items where required),
 'active_writers',(select count(*) from public.pdc_monitor_stage_activation_writers where active and revoked_at is null),
 'active_mailboxes',(select count(*) from public.monitored_mailboxes where active),
 'pilot_enabled',(select enabled from public.pdc_email_monitor_pilot where singleton),
 'monitor_status',(select running_status from public.pdc_email_monitor_status where singleton),
 'gateway_instance_id',(select gateway_instance_id from public.pdc_email_monitor_status where singleton)
 ) evidence;"""
 return _post(f"https://api.supabase.com/v1/projects/{STAGING_REF}/database/query/read-only",sql)[0]["evidence"]

def main():
 a=argparse.ArgumentParser();a.add_argument("--prepare-only",action="store_true");a.add_argument("--apply",action="store_true");a.add_argument("--confirmation");a.add_argument("--receipt",type=Path);args=a.parse_args()
 p,summary=payload();summary["payload_sha256"]=server_hash(p)
 if args.prepare_only:
  print(json.dumps(dict(status="PREPARED",**summary),sort_keys=True));return 0
 if not args.apply or args.confirmation!="CRAIG APPROVED TEMPORARY STAGING IMPORTER ACCESS":raise RuntimeError("PDC_IMPORT_CONFIRMATION_REQUIRED")
 we=env(WEBSITE_ENV);ae=env(AUDITOR_ENV);base=we["PDC_STAGING_SUPABASE_URL"];anon=we["PDC_STAGING_ANON_KEY"]
 if we.get("PDC_STAGING_PROJECT_REF")!=STAGING_REF or PRODUCTION_REF in base or we.get("PDC_PRODUCTION_WRITES_ENABLED","").lower() not in ("false","0","no",""):raise RuntimeError("PDC_IMPORT_TARGET_GUARD")
 admin=auth(base,anon,we["PDC_STAGING_ADMIN_EMAIL"],we["PDC_STAGING_ADMIN_PASSWORD"]);operator=auth(base,anon,we["PDC_STAGING_CONTROLLER_A_EMAIL"],we["PDC_STAGING_CONTROLLER_A_PASSWORD"]);importer=auth(base,anon,ae["PDC_AUDITOR_STAGING_EMAIL"],ae["PDC_AUDITOR_STAGING_PASSWORD"])
 if (admin["role"],operator["role"],importer["role"])!=("administrator","operator","viewer"):raise RuntimeError("PDC_IMPORT_ACTOR_ROLE_MISMATCH")
 before=live_readback()
 if any(before[k] for k in ("vehicles","visible_vehicles","operation_lines","pair_reviews","apply_receipts","active_writers","active_mailboxes")) or before["pilot_enabled"] or before["monitor_status"]!="stopped" or before["gateway_instance_id"] is not None:raise RuntimeError("PDC_IMPORT_PRESTATE_NOT_CONTAINED")
 cleanup=[];primary_error=None
 try:
  role_change(base,admin,importer["email"],"viewer","Craig approved temporary staging workbook importer enrollment")
  rpc(base,admin,"admin_set_pdc_monitor_stage_activation_writer",{"p_monitor_user_id":importer["uid"],"p_active":True,"p_reason":"Craig approved temporary staging operation workbook import"})
  role_change(base,admin,importer["email"],"importer","Craig approved temporary staging operation workbook import")
  importer=auth(base,anon,ae["PDC_AUDITOR_STAGING_EMAIL"],ae["PDC_AUDITOR_STAGING_PASSWORD"])
  preview=rpc(base,importer,"preview_pdc_pmb_retained_workbook",{"p_workbook_sha256":summary["workbook_sha256"],"p_payload_sha256":summary["payload_sha256"],"p_confirmation":"PREVIEW RETAINED PMB WORKBOOK","p_payload":p})
  d=preview["data"]
  expected={"pair_count":166,"operation_count":1488,"applicable_pair_count":0,"approval_required_count":166,"terminal_pair_count":0,"accepted_operation_count":494,"quarantined_operation_count":994}
  if any(d.get(k)!=v for k,v in expected.items()):raise RuntimeError("PDC_PREVIEW_COUNT_MISMATCH:"+json.dumps({k:d.get(k) for k in expected},sort_keys=True))
  pairs=read_pairs(base,admin,d["preview_id"],166)
  if any(x["classification"]!="no_current_stock_manager_override_required" for x in pairs):raise RuntimeError("PDC_PREVIEW_UNEXPECTED_PAIR_CLASSIFICATION")
  if sum(o["disposition"]=="accepted" for x in pairs for o in x["operations"])!=494 or sum(o["disposition"]=="quarantined" for x in pairs for o in x["operations"])!=994:raise RuntimeError("PDC_PREVIEW_OPERATION_DISPOSITION_MISMATCH")
  for x in pairs:
   rpc(base,admin,"approve_pdc_pmb_workbook_pair_exception",{"p_preview_id":d["preview_id"],"p_pair_id":x["pair_id"],"p_workbook_sha256":summary["workbook_sha256"],"p_payload_sha256":summary["payload_sha256"],"p_target_backend_record_id":None,"p_target_vehicle_id":None,"p_expected_vehicle_version":None,"p_reason":"Craig approved exact stock-only staging creation for retained operation workbook"})
  authorization=rpc(base,admin,"authorize_pdc_pmb_workbook_apply",{"p_preview_id":d["preview_id"],"p_workbook_sha256":summary["workbook_sha256"],"p_payload_sha256":summary["payload_sha256"],"p_confirmation":"AUTHORIZE RETAINED PMB WORKBOOK APPLY","p_expected_pair_count":166,"p_expected_operation_count":1488})
  applied=rpc(base,operator,"apply_pdc_pmb_retained_workbook",{"p_preview_id":d["preview_id"],"p_workbook_sha256":summary["workbook_sha256"],"p_payload_sha256":summary["payload_sha256"],"p_confirmation":"APPLY RETAINED PMB WORKBOOK"})
  replay=rpc(base,operator,"apply_pdc_pmb_retained_workbook",{"p_preview_id":d["preview_id"],"p_workbook_sha256":summary["workbook_sha256"],"p_payload_sha256":summary["payload_sha256"],"p_confirmation":"APPLY RETAINED PMB WORKBOOK"})
  if not replay["data"].get("zero_add_replay"):raise RuntimeError("PDC_IMPORT_REPLAY_NOT_ZERO_ADD")
  result={"schema":"pdc-staging-operation-workbook-import-receipt-v1","status":"APPLIED","project_ref":STAGING_REF,**summary,"preview_id":d["preview_id"],"preview_hash":d["preview_hash"],"authorization_id":authorization["data"]["authorization_id"],"receipt_id":applied["data"]["receipt_id"],"receipt_hash":applied["data"]["receipt_hash"],"apply":applied["data"],"replay":replay["data"]}
 except Exception as e:
  primary_error=e;raise
 finally:
  try:
   role_change(base,admin,importer["email"],"viewer","Restore pmb-auditor staging viewer role after approved import");cleanup.append("viewer_restored")
  except Exception as e:cleanup.append("viewer_restore_failed:"+type(e).__name__)
  try:
   rpc(base,admin,"admin_set_pdc_monitor_stage_activation_writer",{"p_monitor_user_id":importer["uid"],"p_active":False,"p_reason":"Revoke temporary staging workbook writer after completed import"});cleanup.append("writer_revoked")
  except Exception as e:cleanup.append("writer_revoke_failed:"+type(e).__name__)
  if primary_error is None and cleanup!=["viewer_restored","writer_revoked"]:raise RuntimeError("PDC_IMPORT_CLEANUP_FAILED:"+json.dumps(cleanup))
 after=live_readback();result["cleanup"]=cleanup;result["readback"]=after
 expected_after={"vehicles":153,"unique_stocks":153,"visible_vehicles":153,"operation_lines":494,"mapped_review_lines":494,"quarantined_review_lines":994,"pair_reviews":166,"pair_approvals":166,"pair_receipts":166,"apply_receipts":1,"active_writers":0,"active_mailboxes":0,"pilot_enabled":False,"monitor_status":"stopped","gateway_instance_id":None}
 if any(after.get(k)!=v for k,v in expected_after.items()):raise RuntimeError("PDC_IMPORT_POSTCONDITION_FAILED:"+json.dumps({k:after.get(k) for k in expected_after},sort_keys=True))
 if args.receipt:
  args.receipt.parent.mkdir(parents=True,exist_ok=True);args.receipt.write_text(json.dumps(result,indent=2,sort_keys=True),encoding="utf-8")
 print(json.dumps({"status":"APPLIED","vehicles":after["vehicles"],"operation_lines":after["operation_lines"],"quarantined_review_lines":after["quarantined_review_lines"],"pair_receipts":after["pair_receipts"],"work_items":after["work_items"],"cleanup":cleanup,"receipt":str(args.receipt.resolve()) if args.receipt else None},sort_keys=True));return 0
if __name__=="__main__":raise SystemExit(main())
