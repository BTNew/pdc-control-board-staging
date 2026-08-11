#!/usr/bin/env python3
"""Build a deterministic Migration-157 payload from a retained PMB open-JC workbook."""
from __future__ import annotations

import argparse
import hashlib
import json
import math
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from scripts.pdc_bulk_workbook_adapter import (
        CONTROL_RE, MAX_OPERATIONS, WorkbookContractError, _hours, infer_work_key,
    )
except ModuleNotFoundError:
    from pdc_bulk_workbook_adapter import (
        CONTROL_RE, MAX_OPERATIONS, WorkbookContractError, _hours, infer_work_key,
    )

OPERATIONS_SHEET_NAME = "Hermes Upload"
OPERATIONS_HEADERS = (
    "Job Card Number", "Stock Number", "Rego", "Operation Line #", "Operation Code",
    "Operation Description", "Estimated Hours", "Hours Provenance", "Match Status",
    "Service Date", "Promised Date",
)
SUMMARY_SHEET_NAME = "Job Card Summary"
SUMMARY_HEADERS = (
    "Job Card Number", "Stock Number", "Rego", "Service Date", "Promised Date",
    "Operation Count", "Match Status", "Schedule Operations",
)

MAX_PAIRS = 600


def _optional(value: Any, label: str, maximum: int, uppercase: bool = False) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    text = str(value).strip()
    if uppercase:
        text = text.upper()
    if CONTROL_RE.search(text) or len(text) > maximum:
        raise WorkbookContractError(f"invalid {label}")
    return text


def _required(value: Any, label: str, maximum: int, uppercase: bool = False) -> str:
    text = _optional(value, label, maximum, uppercase)
    if text is None:
        raise WorkbookContractError(f"missing {label}")
    return text


def _headers(sheet, expected: tuple[str, ...], label: str) -> None:
    actual = tuple(str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1])
    if actual != expected:
        raise WorkbookContractError(f"{label} headers do not match the required contract")


def _reject_formulas(sheet, label: str) -> None:
    if any(cell.data_type == "f" for row in sheet.iter_rows() for cell in row):
        raise WorkbookContractError(f"{label} contains formulas")


def canonical_bytes(payload: list[dict[str, Any]]) -> bytes:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False).encode("utf-8")


def adapt_retained_workbook(path: str | Path) -> tuple[list[dict[str, Any]], bytes, dict[str, Any]]:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() != ".xlsx":
        raise WorkbookContractError("only .xlsx workbooks are accepted")
    workbook_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=False)
    except Exception:
        raise WorkbookContractError("workbook could not be opened safely") from None
    try:
        if OPERATIONS_SHEET_NAME not in wb.sheetnames or SUMMARY_SHEET_NAME not in wb.sheetnames:
            raise WorkbookContractError("required retained workbook sheets are missing")
        upload = wb[OPERATIONS_SHEET_NAME]
        summary = wb[SUMMARY_SHEET_NAME]
        _headers(upload, OPERATIONS_HEADERS, OPERATIONS_SHEET_NAME)
        _headers(summary, SUMMARY_HEADERS, SUMMARY_SHEET_NAME)
        _reject_formulas(upload, OPERATIONS_SHEET_NAME)
        _reject_formulas(summary, SUMMARY_SHEET_NAME)

        pairs: OrderedDict[tuple[str, str | None, str | None], dict[str, Any]] = OrderedDict()
        declared_counts: dict[tuple[str, str | None, str | None], int] = {}
        for values in summary.iter_rows(min_row=2, max_col=len(SUMMARY_HEADERS), values_only=True):
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
                continue
            jc = _required(values[0], "Job Card Number", 60, True)
            stock = _optional(values[1], "Stock Number", 80, True)
            registration = _optional(values[2], "Registration", 40, True)
            match_status = _required(values[6], "Match Status", 30, True)
            schedule = _required(values[7], "Schedule Operations", 3, True)
            if stock is None and registration is None:
                raise WorkbookContractError("summary pair requires Stock or Registration")
            if match_status == "MATCHED" and stock is None:
                raise WorkbookContractError("MATCHED summary pair is missing Stock")
            if match_status == "REGO ONLY" and (stock is not None or registration is None):
                raise WorkbookContractError("REGO ONLY summary pair has invalid identity")
            if match_status not in ("MATCHED", "REGO ONLY") or schedule not in ("YES", "NO"):
                raise WorkbookContractError("invalid summary classification")
            try:
                declared = int(values[5])
            except (TypeError, ValueError):
                raise WorkbookContractError("invalid Operation Count") from None
            if declared < 0 or declared > MAX_OPERATIONS or declared != values[5] or (schedule == "YES") != (declared > 0):
                raise WorkbookContractError("summary operation count/flag disagreement")
            key = (jc, stock, registration)
            if key in pairs:
                raise WorkbookContractError("duplicate retained workbook identity pair")
            if len(pairs) >= MAX_PAIRS:
                raise WorkbookContractError(f"workbook exceeds {MAX_PAIRS} identity pairs")
            pairs[key] = {"job_card_number": jc, "stock_number": stock, "registration": registration, "operations": []}
            declared_counts[key] = declared

        source_line_numbers: dict[tuple[str, str | None, str | None], set[int]] = {key: set() for key in pairs}
        for values in upload.iter_rows(min_row=2, max_col=len(OPERATIONS_HEADERS), values_only=True):
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
                continue
            jc = _required(values[0], "Job Card Number", 60, True)
            stock = _optional(values[1], "Stock Number", 80, True)
            registration = _optional(values[2], "Registration", 40, True)
            match_status = _required(values[8], "Match Status", 30, True)
            if match_status == "MATCHED" and stock is None:
                raise WorkbookContractError("MATCHED operation row is missing Stock")
            if match_status == "REGO ONLY" and (stock is not None or registration is None):
                raise WorkbookContractError("REGO ONLY operation row has invalid identity")
            key = (jc, stock, registration)
            if key not in pairs:
                raise WorkbookContractError("operation row has no exact summary identity pair")
            try:
                source_line = int(values[3])
            except (TypeError, ValueError):
                raise WorkbookContractError("invalid Operation Line #") from None
            if source_line < 1 or source_line != values[3] or source_line in source_line_numbers[key]:
                raise WorkbookContractError("invalid or duplicate Operation Line #")
            source_line_numbers[key].add(source_line)
            description = _required(values[5], "Operation Description", 180)
            hours, source = _hours(values[6])
            if hours is None:
                hours, source = 1, "ai_estimate"
            operations = pairs[key]["operations"]
            operations.append({
                "operation_no": f"OP{len(operations) + 1}",
                "work_key": infer_work_key(description),
                "description": description,
                "estimated_hours": hours,
                "estimated_hours_source": source,
            })

        for key, pair in pairs.items():
            if len(pair["operations"]) != declared_counts[key]:
                raise WorkbookContractError("summary Operation Count does not match retained operation rows")
        payload = [{"pair_no": number, **pair} for number, pair in enumerate(pairs.values(), 1)]
    finally:
        wb.close()

    canonical = canonical_bytes(payload)
    operations = [op for pair in payload for op in pair["operations"]]
    hours = [float(op["estimated_hours"]) for op in operations]
    evidence = {
        "code": "retained_pmb_workbook_adapted",
        "workbook_sha256": workbook_sha,
        "canonical_json_sha256": hashlib.sha256(canonical).hexdigest(),
        "pair_count": len(payload),
        "stock_pair_count": sum(1 for pair in payload if pair["stock_number"] is not None),
        "registration_only_pair_count": sum(1 for pair in payload if pair["stock_number"] is None),
        "operation_count": len(operations),
        "source_hours_count": sum(1 for op in operations if op["estimated_hours_source"] == "job_card"),
        "sixty_minute_fallback_count": sum(1 for op in operations if op["estimated_hours_source"] == "ai_estimate"),
        "estimated_hours_total": round(math.fsum(hours), 2),
        "stock_only_rows": sum(1 for pair in payload if not pair["operations"]),
        "max_operations_per_pair": max(len(pair["operations"]) for pair in payload),
        "stage_counts": {key: sum(1 for op in operations if op["work_key"] == key) for key in sorted({op["work_key"] for op in operations})},
    }
    return payload, canonical, evidence


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--payload-out", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        payload, canonical, evidence = adapt_retained_workbook(args.workbook)
        args.payload_out.write_bytes(canonical)
        print(json.dumps(evidence, sort_keys=True, separators=(",", ":")))
        return 0
    except (OSError, WorkbookContractError) as exc:
        print(json.dumps({"code": "retained_workbook_adapter_failed", "error": str(exc)}, sort_keys=True), file=__import__("sys").stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
