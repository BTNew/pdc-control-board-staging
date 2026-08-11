#!/usr/bin/env python3
"""Convert a retained JC/Stock XLSX into a deterministic, reviewable payload.

The workbook's business rows are never printed.  The CLI writes the canonical
payload only to an explicitly selected file and emits sanitized aggregate
identity evidence on stdout.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from collections import OrderedDict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET_NAME = "Hermes Import"
HEADERS = ("JC Number", "Stock Number", "Operation", "Estimated Hours")
CONTROL_RE = re.compile(r"[\x00-\x1f\x7f-\x9f]")
MAX_ROWS = 500
MAX_OPERATIONS = 100
STAGE_MAPPING_POLICY = "pmb-workshop-stages-v2"
PLACEHOLDER_OPERATION = "no operation available"


def _normalized_words(value: str) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value.casefold()).split())


def is_placeholder_operation(description: str) -> bool:
    return _normalized_words(description) == PLACEHOLDER_OPERATION


def _contains_any(text: str, phrases: tuple[str, ...]) -> bool:
    return any(phrase in text for phrase in phrases)


def infer_work_key(description: str) -> str:
    """Map a real workbook operation to one deterministic PMB workshop stage."""
    text = _normalized_words(description)
    if not text or text == PLACEHOLDER_OPERATION:
        raise WorkbookContractError("placeholder operation has no workshop stage")
    # Craig-verified routing exceptions must run before broad words such as
    # tint, headlamp, loose, or wiring can influence the department.
    if _contains_any(text, (
        "first aid kit", "fire extinguisher", "safety triangle", "bonnet protector",
        "headlamp cover", "headlight cover", "seat cover", "pre delivery",
        "predelivery", "pdi", "recovery point", "tow bar",
    )):
        return "fitting"
    if _contains_any(text, (
        "long range tank", "long range fuel tank", "replacement fuel tank",
        "arb frontier tank", "arb frontier fuel tank",
    )):
        return "hoist"
    if _contains_any(text, ("pte tray at cost", "sublet", "outsourc")):
        return "sublet"
    if "pit" in text.split() or _contains_any(text, ("pit and weigh", "pit weigh", "weighbridge", "fill with fuel", "fuel fill")):
        return "pitInspection"
    if "tint" in text:
        return "tint"
    if _contains_any(text, ("4x4 bus", "bus 4x4", "4x4 conversion", "four by four conversion")):
        return "bus4x4"
    if _contains_any(text, ("gvm", "suspension", "shock absorber", "leaf spring", "coil spring", "lift kit")) or re.search(r"\bome\b", text):
        return "hoist"
    if _contains_any(text, ("tyre", "tire", "wheel alignment", "spare wheel", "spare rim", "steel rim", "alloy rim", "wheel nut indicator")) or re.search(r"\brim\b", text):
        return "tyre"
    if _contains_any(text, ("left loose", "loose in vehicle", "loose in car", "supply only", "do not fit")):
        return "PARTS"
    if _contains_any(text, (
        "battery", "isolator", "minebar", "lightbar", "headlamp", "lamp", "beacon",
        "reverse buzzer", "brake controller", "uhf", "gme", "radio", "aerial", "antenna",
        "wiring", "wired", "switch", "camera", "dashcam", "plug", "usb", "electrical",
        "electric", "sounder", "siren", "horn", "inverter", "solar", "redarc", "strobe",
        "narva", " led ", "alarm", "tracker", " gps ", "ignition", "hand brake",
    )):
        return "electrical"
    if _contains_any(text, ("fabricat", "weld", "custom bracket", "tray modification", "chassis modification")):
        return "fabrication"
    return "fitting"


class WorkbookContractError(ValueError):
    """Safe, business-data-free workbook validation failure."""


@dataclass(frozen=True)
class AdaptedWorkbook:
    payload: list[dict[str, Any]]
    canonical_json: bytes
    evidence: dict[str, Any]


def _text(value: Any, label: str, maximum: int) -> str:
    if value is None:
        raise WorkbookContractError(f"missing {label}")
    text = str(value).strip()
    if not text:
        raise WorkbookContractError(f"missing {label}")
    if CONTROL_RE.search(text):
        raise WorkbookContractError(f"control character in {label}")
    if len(text) > maximum:
        raise WorkbookContractError(f"{label} exceeds {maximum} characters")
    return text


def _hours(value: Any) -> tuple[int | float | None, str | None]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, None
    if isinstance(value, bool):
        raise WorkbookContractError("invalid Schedule Hrs")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise WorkbookContractError("invalid Schedule Hrs") from None
    if not number.is_finite() or number < 0 or number > Decimal("999.99"):
        raise WorkbookContractError("Schedule Hrs outside 0..999.99")
    if number.as_tuple().exponent < -2:
        raise WorkbookContractError("Schedule Hrs has more than two decimal places")
    number = number.quantize(Decimal("0.01")).normalize()
    wire: int | float = int(number) if number == number.to_integral() else float(number)
    return wire, "job_card"


def canonical_payload_bytes(payload: list[dict[str, Any]]) -> bytes:
    """Return UTF-8 canonical JSON (stable key order, no insignificant space)."""
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False).encode("utf-8")


def adapt_workbook(path: str | Path, stage_mapping_policy: str | None = None) -> AdaptedWorkbook:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() != ".xlsx":
        raise WorkbookContractError("only .xlsx workbooks are accepted")
    workbook_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    except Exception:
        raise WorkbookContractError("workbook could not be opened safely") from None
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise WorkbookContractError(f"required worksheet {SHEET_NAME!r} is missing")
        sheet = workbook[SHEET_NAME]
        header_cells = list(sheet[4])
        header_values = [cell.value for cell in header_cells]
        populated_headers = [value for value in header_values if value is not None and str(value).strip()]
        normalized = tuple(str(value).strip() for value in populated_headers)
        if len(normalized) != len(set(normalized)):
            raise WorkbookContractError("duplicate header in row 4")
        padded = header_values + [None] * max(0, 4 - len(header_values))
        positioned = tuple(str(value).strip() if value is not None else "" for value in padded[:4])
        has_extra = any(value is not None and str(value).strip() for value in header_values[4:])
        if positioned != HEADERS or has_extra:
            raise WorkbookContractError("row 4 headers do not match the required contract")

        groups: OrderedDict[tuple[str, str], list[dict[str, Any]]] = OrderedDict()
        seen_operations: dict[tuple[str, str], set[tuple[str, int | float | None]]] = {}
        for values in sheet.iter_rows(min_row=5, max_col=4, values_only=True):
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
                continue
            jc = _text(values[0], "JC", 60)
            stock = _text(values[1], "Stock", 80)
            description = _text(values[2], "Operation / Kit", 180)
            key = (jc, stock)
            if key not in groups:
                if len(groups) >= MAX_ROWS:
                    raise WorkbookContractError(f"workbook exceeds {MAX_ROWS} JC/Stock groups")
                groups[key] = []
                seen_operations[key] = set()
            if stage_mapping_policy == STAGE_MAPPING_POLICY and is_placeholder_operation(description):
                continue
            if stage_mapping_policy not in (None, STAGE_MAPPING_POLICY):
                raise WorkbookContractError("unsupported workshop-stage mapping policy")
            hours, source = _hours(values[3])
            duplicate_key = (description, hours)
            if duplicate_key in seen_operations[key]:
                raise WorkbookContractError("duplicate operation within a JC/Stock group")
            seen_operations[key].add(duplicate_key)
            operations = groups[key]
            if len(operations) >= MAX_OPERATIONS:
                raise WorkbookContractError(f"JC/Stock group exceeds {MAX_OPERATIONS} operations")
            operations.append({
                "operation_no": f"OP{len(operations) + 1}",
                "work_key": infer_work_key(description) if stage_mapping_policy else None,
                "description": description,
                "estimated_hours": hours,
                "estimated_hours_source": source,
            })
    finally:
        workbook.close()

    if not groups:
        raise WorkbookContractError("workbook contains no data rows")
    payload = [
        {"row_no": row_no, "job_card_number": jc, "stock_number": stock, "operations": operations}
        for row_no, ((jc, stock), operations) in enumerate(groups.items(), 1)
    ]
    canonical = canonical_payload_bytes(payload)
    operation_count = sum(len(row["operations"]) for row in payload)
    hours_values = [op["estimated_hours"] for row in payload for op in row["operations"] if op["estimated_hours"] is not None]
    evidence = {
        "code": "workbook_adapted",
        "workbook_sha256": workbook_sha,
        "payload_sha256": hashlib.sha256(canonical).hexdigest(),
        "jc_stock_pair_count": len(payload),
        "operation_count": operation_count,
        "estimated_hours_count": len(hours_values),
        "missing_hours_count": operation_count - len(hours_values),
        "estimated_hours_total": round(math.fsum(float(value) for value in hours_values), 2),
        "max_operations_per_pair": max(len(row["operations"]) for row in payload),
        "stage_mapping_policy": stage_mapping_policy,
        "stage_counts": {
            key: sum(1 for row in payload for op in row["operations"] if op["work_key"] == key)
            for key in sorted({op["work_key"] for row in payload for op in row["operations"] if op["work_key"] is not None})
        },
        "stock_only_rows": sum(1 for row in payload if not row["operations"]),
    }
    return AdaptedWorkbook(payload, canonical, evidence)


def assert_expected(evidence: dict[str, Any], args: argparse.Namespace) -> None:
    checks = {
        "jc_stock_pair_count": args.expect_pairs,
        "operation_count": args.expect_operations,
        "estimated_hours_count": args.expect_hours_count,
        "missing_hours_count": args.expect_missing_hours,
        "estimated_hours_total": args.expect_hours_total,
        "max_operations_per_pair": args.expect_max_operations,
    }
    for field, expected in checks.items():
        if expected is not None and evidence[field] != expected:
            raise WorkbookContractError(f"aggregate assertion failed for {field}")


def add_assertion_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--expect-pairs", type=int)
    parser.add_argument("--expect-operations", type=int)
    parser.add_argument("--expect-hours-count", type=int)
    parser.add_argument("--expect-missing-hours", type=int)
    parser.add_argument("--expect-hours-total", type=float)
    parser.add_argument("--expect-max-operations", type=int)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--payload-out", type=Path, required=True, help="exclusive output path for canonical business payload")
    parser.add_argument("--stage-map-policy", choices=[STAGE_MAPPING_POLICY])
    add_assertion_arguments(parser)
    args = parser.parse_args(argv)
    try:
        adapted = adapt_workbook(args.workbook, stage_mapping_policy=args.stage_map_policy)
        assert_expected(adapted.evidence, args)
        args.payload_out.write_bytes(adapted.canonical_json)
        print(json.dumps(adapted.evidence, sort_keys=True, separators=(",", ":")))
        return 0
    except (OSError, WorkbookContractError) as exc:
        print(json.dumps({"code": "workbook_adapter_failed", "error": str(exc)}, sort_keys=True), file=__import__("sys").stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
