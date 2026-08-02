from __future__ import annotations

import argparse
import hashlib
import io
import json
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from urllib.error import HTTPError
from unittest.mock import patch

from openpyxl import Workbook

from scripts.pdc_bulk_workbook_adapter import WorkbookContractError, adapt_workbook, main as adapter_main
from scripts.run_pdc_bulk_workbook_staging import EXPECTED_URL, RunnerError, _post, execute


HEADERS = ("JC", "Stock", "Operation / Kit", "Schedule Hrs")
ADMIN_ENV = {
    "PDC_STAGING_SUPABASE_URL": EXPECTED_URL,
    "PDC_STAGING_ANON_KEY": "anon-public-value",
    "PDC_STAGING_ADMIN_EMAIL": "admin@example.invalid",
    "PDC_STAGING_ADMIN_PASSWORD": "admin-secret",
    "PDC_STAGING_VIEWER_EMAIL": "viewer@example.invalid",
    "PDC_STAGING_VIEWER_PASSWORD": "viewer-secret",
    "PDC_STAGING_SERVICE_ROLE_KEY": "service-secret",
}


def make_workbook(path: Path, rows, headers=HEADERS):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hermes Import"
    for column, value in enumerate(headers, 1):
        sheet.cell(4, column, value)
    for row_no, values in enumerate(rows, 5):
        for column, value in enumerate(values, 1):
            sheet.cell(row_no, column, value)
    workbook.save(path)


def args(path: Path, *, apply=False, preview_id=None, workbook_sha=None, payload_sha=None):
    return argparse.Namespace(
        workbook=path,
        apply=apply,
        confirm_preview_id=preview_id,
        confirm_workbook_sha256=workbook_sha,
        confirm_payload_sha256=payload_sha,
        expect_pairs=None,
        expect_operations=None,
        expect_hours_count=None,
        expect_missing_hours=None,
        expect_hours_total=None,
    )


class AdapterTests(unittest.TestCase):
    def test_grouping_aggregates_nulls_and_deterministic_hash(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            make_workbook(path, [
                (" JC-2 ", " ST-2 ", " First operation ", 1.25),
                ("JC-1", "ST-1", "Other", None),
                ("JC-2", "ST-2", "Second", 0),
            ])
            first, second = adapt_workbook(path), adapt_workbook(path)
            self.assertEqual(first.canonical_json, second.canonical_json)
            self.assertEqual(first.evidence["payload_sha256"], hashlib.sha256(first.canonical_json).hexdigest())
            self.assertEqual(first.evidence["jc_stock_pair_count"], 2)
            self.assertEqual(first.evidence["operation_count"], 3)
            self.assertEqual(first.evidence["estimated_hours_count"], 2)
            self.assertEqual(first.evidence["missing_hours_count"], 1)
            self.assertEqual(first.evidence["estimated_hours_total"], 1.25)
            self.assertEqual(first.payload[0]["row_no"], 1)
            self.assertEqual(first.payload[0]["operations"][0]["operation_no"], "OP001")
            self.assertEqual(first.payload[0]["operations"][1]["operation_no"], "OP002")
            self.assertIsNone(first.payload[1]["operations"][0]["work_key"])
            self.assertIsNone(first.payload[1]["operations"][0]["estimated_hours"])
            self.assertIsNone(first.payload[1]["operations"][0]["estimated_hours_source"])
            self.assertEqual(first.payload[0]["operations"][0]["estimated_hours_source"], "job_card")

    def test_max_operations_and_bounds(self):
        with tempfile.TemporaryDirectory() as directory:
            accepted = Path(directory) / "accepted.xlsx"
            make_workbook(accepted, [("JC", "STOCK", f"operation {i}", 999.99) for i in range(100)])
            self.assertEqual(adapt_workbook(accepted).evidence["operation_count"], 100)
            rejected = Path(directory) / "rejected.xlsx"
            make_workbook(rejected, [("JC", "STOCK", f"operation {i}", 1) for i in range(101)])
            with self.assertRaisesRegex(WorkbookContractError, "exceeds 100"):
                adapt_workbook(rejected)
            bad_hours = Path(directory) / "bad-hours.xlsx"
            make_workbook(bad_hours, [("JC", "STOCK", "operation", 1000)])
            with self.assertRaisesRegex(WorkbookContractError, "outside 0..999.99"):
                adapt_workbook(bad_hours)

    def test_cli_writes_payload_only_to_file_and_checks_aggregates(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "synthetic.xlsx"
            payload_path = Path(directory) / "payload.json"
            make_workbook(path, [("PRIVATE-JC", "PRIVATE-STOCK", "private operation", 1)])
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                return_code = adapter_main([
                    str(path), "--payload-out", str(payload_path),
                    "--expect-pairs", "1", "--expect-operations", "1",
                    "--expect-hours-count", "1", "--expect-missing-hours", "0",
                    "--expect-hours-total", "1.0",
                ])
            self.assertEqual(return_code, 0)
            evidence = json.loads(stdout.getvalue())
            self.assertEqual(evidence["code"], "workbook_adapted")
            self.assertNotIn("PRIVATE", stdout.getvalue())
            self.assertIn("PRIVATE-JC", payload_path.read_text("utf-8"))
            self.assertEqual(hashlib.sha256(payload_path.read_bytes()).hexdigest(), evidence["payload_sha256"])

    def test_header_control_and_duplicate_rejection(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            bad_header = root / "header.xlsx"
            make_workbook(bad_header, [("JC", "ST", "op", 1)], ("JC", "Stock", "JC", "Schedule Hrs"))
            with self.assertRaisesRegex(WorkbookContractError, "duplicate header"):
                adapt_workbook(bad_header)
            control = root / "control.xlsx"
            # C1 control U+0085 is XML-valid (unlike C0 controls) and therefore
            # reaches the adapter through openpyxl for an executable rejection.
            make_workbook(control, [("JC", "ST", "unsafe\x85operation", 1)])
            with self.assertRaisesRegex(WorkbookContractError, "control character"):
                adapt_workbook(control)
            duplicate = root / "duplicate.xlsx"
            make_workbook(duplicate, [("JC", "ST", "same", 1), ("JC", "ST", "same", 1)])
            with self.assertRaisesRegex(WorkbookContractError, "duplicate operation"):
                adapt_workbook(duplicate)


class RunnerTests(unittest.TestCase):
    def workbook(self, root: str) -> tuple[Path, object]:
        path = Path(root) / "runner.xlsx"
        make_workbook(path, [("JC", "ST", "operation", 2.5)])
        return path, adapt_workbook(path)

    def transport(self, adapted, calls, include_apply=False):
        preview_id = "12345678-1234-4123-8123-123456789abc"

        def post(url, key, path, body, token=None):
            calls.append((url, key, path, body, token))
            if path.startswith("/auth/"):
                return {"access_token": "access-secret"}
            if path.endswith("preview_pdc_bulk_jc_stock_workbook"):
                return {"ok": True, "code": "preview_ready", "data": {
                    "preview_id": preview_id,
                    "workbook_sha256": adapted.evidence["workbook_sha256"],
                    "payload_sha256": adapted.evidence["payload_sha256"],
                }}
            if include_apply and path.endswith("apply_pdc_bulk_jc_stock_workbook"):
                return {"ok": True, "code": "applied", "data": {
                    "receipt_hash": "a" * 64, "row_count": 1, "quarantine_count": 0,
                    "vehicles_added": 1, "operation_lines_added": 1, "estimated_hours_added": 1,
                }}
            raise AssertionError(path)
        return preview_id, post

    def test_staging_guard(self):
        with tempfile.TemporaryDirectory() as directory:
            path, _ = self.workbook(directory)
            bad = dict(ADMIN_ENV, PDC_STAGING_SUPABASE_URL="https://example.invalid")
            with self.assertRaisesRegex(RunnerError, "target guard"):
                execute(args(path), bad, lambda *unused: {})

    def test_admin_env_only_and_preview_default(self):
        with tempfile.TemporaryDirectory() as directory:
            path, adapted = self.workbook(directory)
            calls = []
            _, post = self.transport(adapted, calls)
            result = execute(args(path), ADMIN_ENV, post)
            self.assertFalse(result["apply_performed"])
            self.assertEqual([call[2] for call in calls], [
                "/auth/v1/token?grant_type=password",
                "/rest/v1/rpc/preview_pdc_bulk_jc_stock_workbook",
            ])
            self.assertEqual(calls[0][3], {"email": ADMIN_ENV["PDC_STAGING_ADMIN_EMAIL"], "password": ADMIN_ENV["PDC_STAGING_ADMIN_PASSWORD"]})
            serialized_calls = repr(calls)
            self.assertNotIn("viewer-secret", serialized_calls)
            self.assertNotIn("service-secret", serialized_calls)

    def test_apply_requires_three_exact_confirmations(self):
        with tempfile.TemporaryDirectory() as directory:
            path, adapted = self.workbook(directory)
            calls = []
            preview_id, post = self.transport(adapted, calls, include_apply=True)
            incomplete = args(path, apply=True, preview_id=preview_id, workbook_sha=adapted.evidence["workbook_sha256"])
            with self.assertRaisesRegex(RunnerError, "confirmations"):
                execute(incomplete, ADMIN_ENV, post)
            self.assertEqual(calls, [])
            calls.clear()
            exact = args(path, apply=True, preview_id=preview_id, workbook_sha=adapted.evidence["workbook_sha256"], payload_sha=adapted.evidence["payload_sha256"])
            result = execute(exact, ADMIN_ENV, post)
            self.assertTrue(result["apply_performed"])
            self.assertEqual(result["code"], "applied")
            self.assertEqual(calls[-1][3]["p_preview_id"], preview_id)

    def test_http_error_redacts_response_and_secrets(self):
        secret = b'{"message":"database detail","token":"super-secret"}'
        error = HTTPError(EXPECTED_URL, 403, "forbidden", {}, None)
        error.read = lambda: secret
        with patch("scripts.run_pdc_bulk_workbook_staging.urlopen", side_effect=error):
            with self.assertRaises(RunnerError) as caught:
                _post(EXPECTED_URL, "anon-secret", "/rest/v1/rpc/test", {"password": "admin-secret"}, "token-secret")
        message = str(caught.exception)
        self.assertIn("403", message)
        for value in ("database detail", "super-secret", "anon-secret", "admin-secret", "token-secret"):
            self.assertNotIn(value, message)


if __name__ == "__main__":
    unittest.main()
