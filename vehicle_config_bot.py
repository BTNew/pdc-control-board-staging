"""Credential-free, fail-closed processing core for Vehicle Config workbooks.

The module deliberately contains no network, bot-token, or Supabase integration.  A
trusted adapter may turn a reviewed :class:`ChangeSet` into a file update, but this
module always validates the resulting document before replacing the destination.
"""
from __future__ import annotations

import csv
import io
import os
import re
import tempfile
from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Sequence

ALLOWED_FIELDS = frozenset({"Hidden", "Cost", "Sell"})
PRIVILEGED_COMMANDS = frozenset({"remember", "correct", "disable", "undo"})
READ_COMMANDS = frozenset({"review", "apply", "explain", "show unresolved"})


class VehicleConfigError(Exception):
    """Base error for a rejected Vehicle Config operation."""


class UnsupportedFormatError(VehicleConfigError):
    pass


class WorkbookDependencyError(VehicleConfigError):
    pass


class ValidationError(VehicleConfigError):
    pass


class AuthorizationError(VehicleConfigError):
    pass


@dataclass(frozen=True)
class CellChange:
    """One proposed edit, addressed by sheet, one-based data row, and header."""

    row: int
    field: str
    value: Any
    sheet: str | None = None


@dataclass(frozen=True)
class ParsedCommand:
    action: str
    argument: str = ""
    requires_authorized_identity: bool = False


@dataclass(frozen=True)
class PricingResult:
    cost_ex_gst: Decimal
    sell_inc_gst: Decimal
    solis_applied: bool
    hilux_gvm_inc_gst: Decimal


FREIGHT_COST_EX_GST = {
    "kununurra": Decimal("830.00"),
    "halls creek": Decimal("830.00"),
    "fitzroy crossing": Decimal("830.00"),
    "derby": Decimal("600.00"),
}


def calculate_pmb_sell(pmb_cost_ex_gst: Any) -> Decimal:
    """PMB sell is explicit cost +10% margin, then 10% GST."""
    cost = _money(pmb_cost_ex_gst, "pmb_cost_ex_gst")
    return (cost * Decimal("1.10") * Decimal("1.10")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_freight(destination: str) -> tuple[Decimal, Decimal]:
    key = " ".join(str(destination).strip().lower().split())
    if key not in FREIGHT_COST_EX_GST:
        raise ValidationError("freight destination has no approved price evidence")
    cost = FREIGHT_COST_EX_GST[key]
    sell = (cost * Decimal("1.20") * Decimal("1.10")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return cost, sell


def approved_pd_pricing() -> tuple[Decimal, Decimal]:
    return Decimal("300.00"), Decimal("1995.00")


def _money(value: Any, name: str) -> Decimal:
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValidationError(f"{name} must be a decimal amount") from exc
    if not result.is_finite() or result < 0:
        raise ValidationError(f"{name} must be a finite, non-negative amount")
    return result.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_pricing(
    *,
    cost_ex_gst: Any,
    sell_inc_gst: Any,
    solis: bool = False,
    hilux_gvm_inc_gst: Any = 0,
    hilux_gvm_approved: bool = False,
) -> PricingResult:
    """Use only supplied ARB/base evidence; Solis is the sole automatic add-on.

    Cost input is ex GST and Sell input is inclusive GST. A Hilux GVM amount is
    accepted only when the caller also supplies explicit approval evidence.
    """

    cost = _money(cost_ex_gst, "cost_ex_gst")
    sell = _money(sell_inc_gst, "sell_inc_gst")
    gvm = _money(hilux_gvm_inc_gst, "hilux_gvm_inc_gst")
    if gvm and not hilux_gvm_approved:
        raise ValidationError("Hilux GVM pricing requires explicit approval")
    solis_charge = Decimal("150.00") if solis else Decimal("0.00")
    return PricingResult(
        cost_ex_gst=cost,
        sell_inc_gst=(sell + solis_charge + gvm).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        solis_applied=bool(solis),
        hilux_gvm_inc_gst=gvm,
    )


def parse_command(text: str, *, authorized_identity: str | None = None) -> ParsedCommand:
    """Parse the small command surface, requiring identity for learning/mutation stubs."""

    normalized = " ".join(text.strip().split())
    match = re.fullmatch(
        r"(?i)(show\s+unresolved|review|apply|explain|remember|correct|disable|undo)(?:\s+(.*))?",
        normalized,
    )
    if not match:
        raise ValidationError("unsupported command")
    action = match.group(1).lower()
    action = "show unresolved" if action.startswith("show") else action
    argument = match.group(2) or ""
    privileged = action in PRIVILEGED_COMMANDS
    if privileged and not (authorized_identity and authorized_identity.strip()):
        raise AuthorizationError(f"{action.title()} requires an authorized identity")
    return ParsedCommand(action, argument, privileged)


def execute_privileged_stub(command: ParsedCommand) -> None:
    """Fail closed: persistence/learning is intentionally outside this foundation."""

    if command.action not in PRIVILEGED_COMMANDS:
        raise ValidationError("not a privileged command")
    raise VehicleConfigError(
        f"{command.action.title()} is not configured; an authorized credential-backed adapter is required"
    )


def _require_changes(changes: Iterable[CellChange]) -> tuple[CellChange, ...]:
    result = tuple(changes)
    seen: set[tuple[str | None, int, str]] = set()
    for change in result:
        if change.field not in ALLOWED_FIELDS:
            raise ValidationError(f"field {change.field!r} is not editable")
        if isinstance(change.row, bool) or not isinstance(change.row, int) or change.row < 1:
            raise ValidationError("row must be a one-based positive integer")
        key = (change.sheet, change.row, change.field)
        if key in seen:
            raise ValidationError(f"duplicate change target: {key}")
        seen.add(key)
    return result


def apply_file(source: str | os.PathLike[str], destination: str | os.PathLike[str], changes: Iterable[CellChange]) -> None:
    """Apply edits through a temporary file and a strict before/after gate."""

    src, dst = Path(source), Path(destination)
    if not src.is_file():
        raise ValidationError(f"source does not exist: {src}")
    requested = _require_changes(changes)
    suffix = src.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise UnsupportedFormatError("only CSV and XLSX are supported")
    if dst.resolve() != src.resolve() and dst.exists():
        raise ValidationError("destination already exists")
    dst.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=".vehicle-config-", suffix=suffix, dir=dst.parent)
    os.close(fd)
    temp = Path(temp_name)
    try:
        if suffix == ".csv":
            _apply_csv(src, temp, requested)
            _validate_csv(src, temp, requested)
        else:
            _apply_xlsx(src, temp, requested)
            _validate_xlsx(src, temp, requested)
        os.replace(temp, dst)
    except Exception:
        temp.unlink(missing_ok=True)
        raise


def _csv_document(path: Path) -> tuple[str, csv.Dialect, list[list[str]]]:
    raw = path.read_bytes()
    encoding = "utf-8-sig" if raw.startswith(b"\xef\xbb\xbf") else "utf-8"
    text = raw.decode(encoding)
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return encoding, dialect, list(csv.reader(io.StringIO(text, newline=""), dialect))


def _apply_csv(src: Path, out: Path, changes: Sequence[CellChange]) -> None:
    encoding, dialect, rows = _csv_document(src)
    if not rows:
        raise ValidationError("CSV has no header row")
    header = rows[0]
    if len(header) != len(set(header)):
        raise ValidationError("CSV headers must be unique")
    indexes = {name: index for index, name in enumerate(header)}
    for change in changes:
        if change.sheet is not None:
            raise ValidationError("CSV changes cannot specify a sheet")
        if change.field not in indexes:
            raise ValidationError(f"missing required header: {change.field}")
        physical_row = change.row
        if physical_row >= len(rows):
            raise ValidationError(f"data row {change.row} does not exist")
        if len(rows[physical_row]) != len(header):
            raise ValidationError(f"data row {change.row} has a different column count")
        rows[physical_row][indexes[change.field]] = str(change.value)
    newline = "\r\n" if b"\r\n" in src.read_bytes() else "\n"
    with out.open("w", encoding=encoding, newline="") as handle:
        writer = csv.writer(
            handle,
            delimiter=dialect.delimiter,
            quotechar=dialect.quotechar,
            escapechar=dialect.escapechar,
            doublequote=dialect.doublequote,
            quoting=dialect.quoting,
            lineterminator=newline,
        )
        writer.writerows(rows)


def _validate_csv(before: Path, after: Path, changes: Sequence[CellChange]) -> None:
    _, _, old = _csv_document(before)
    _, _, new = _csv_document(after)
    if len(old) != len(new) or any(len(a) != len(b) for a, b in zip(old, new)):
        raise ValidationError("CSV structure changed")
    if not old or old[0] != new[0]:
        raise ValidationError("CSV headers or column order changed")
    allowed = {(change.row, old[0].index(change.field)) for change in changes}
    for row_index, (old_row, new_row) in enumerate(zip(old, new)):
        for col_index, (old_value, new_value) in enumerate(zip(old_row, new_row)):
            if old_value != new_value and (row_index, col_index) not in allowed:
                raise ValidationError(f"unauthorized CSV change at row {row_index}, column {col_index + 1}")


def _openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise WorkbookDependencyError("XLSX support requires the optional 'openpyxl' package") from exc
    return openpyxl


def _sheet_headers(sheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        if cell.value is not None:
            if not isinstance(cell.value, str) or cell.value in headers:
                raise ValidationError(f"sheet {sheet.title!r} has invalid or duplicate headers")
            headers[cell.value] = cell.column
    return headers


def _apply_xlsx(src: Path, out: Path, changes: Sequence[CellChange]) -> None:
    openpyxl = _openpyxl()
    keep_vba = src.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(src, data_only=False, keep_vba=keep_vba)
    try:
        for change in changes:
            if change.sheet is None:
                if len(workbook.sheetnames) != 1:
                    raise ValidationError("sheet is required for a multi-sheet workbook")
                sheet = workbook[workbook.sheetnames[0]]
            elif change.sheet not in workbook.sheetnames:
                raise ValidationError(f"sheet does not exist: {change.sheet}")
            else:
                sheet = workbook[change.sheet]
            headers = _sheet_headers(sheet)
            if change.field not in headers:
                raise ValidationError(f"missing required header: {change.field}")
            physical_row = change.row + 1
            if physical_row > sheet.max_row:
                raise ValidationError(f"data row {change.row} does not exist")
            sheet.cell(physical_row, headers[change.field]).value = change.value
        workbook.save(out)
    finally:
        workbook.close()


def _style_signature(cell: Any) -> tuple[Any, ...]:
    return (
        cell.style_id,
        cell.number_format,
        copy(cell.alignment),
        copy(cell.protection),
    )


def _validate_xlsx(before: Path, after: Path, changes: Sequence[CellChange]) -> None:
    openpyxl = _openpyxl()
    old_book = openpyxl.load_workbook(before, data_only=False)
    new_book = openpyxl.load_workbook(after, data_only=False)
    try:
        if old_book.sheetnames != new_book.sheetnames:
            raise ValidationError("sheet names or order changed")
        allowed: set[tuple[str, int, int]] = set()
        for change in changes:
            sheet_name = change.sheet or old_book.sheetnames[0]
            headers = _sheet_headers(old_book[sheet_name])
            allowed.add((sheet_name, change.row + 1, headers[change.field]))
        for sheet_name in old_book.sheetnames:
            old, new = old_book[sheet_name], new_book[sheet_name]
            if (old.max_row, old.max_column) != (new.max_row, new.max_column):
                raise ValidationError(f"sheet structure changed: {sheet_name}")
            if list(old.merged_cells.ranges) != list(new.merged_cells.ranges):
                raise ValidationError(f"merged cells changed: {sheet_name}")
            for row in range(1, old.max_row + 1):
                for column in range(1, old.max_column + 1):
                    a, b = old.cell(row, column), new.cell(row, column)
                    target = (sheet_name, row, column)
                    if a.value != b.value and target not in allowed:
                        raise ValidationError(f"unauthorized XLSX change at {sheet_name}!{b.coordinate}")
                    if _style_signature(a) != _style_signature(b):
                        raise ValidationError(f"formatting changed at {sheet_name}!{b.coordinate}")
            if old.freeze_panes != new.freeze_panes or old.sheet_format != new.sheet_format:
                raise ValidationError(f"sheet formatting changed: {sheet_name}")
            for key, dimension in old.column_dimensions.items():
                if key not in new.column_dimensions or dimension.width != new.column_dimensions[key].width:
                    raise ValidationError(f"column formatting changed: {sheet_name}!{key}")
            for key, dimension in old.row_dimensions.items():
                if key not in new.row_dimensions or dimension.height != new.row_dimensions[key].height:
                    raise ValidationError(f"row formatting changed: {sheet_name}!{key}")
    finally:
        old_book.close()
        new_book.close()


def review_file(path: str | os.PathLike[str], changes: Iterable[CellChange]) -> tuple[CellChange, ...]:
    """Validate proposed targets without writing a file; suitable for Review adapters."""

    requested = _require_changes(changes)
    src = Path(path)
    fd, name = tempfile.mkstemp(suffix=src.suffix)
    os.close(fd)
    temp = Path(name)
    temp.unlink(missing_ok=True)
    try:
        apply_file(src, temp, requested)
    finally:
        temp.unlink(missing_ok=True)
    return requested
